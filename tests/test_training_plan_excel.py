from copy import deepcopy
import io

import app as app_module
import openpyxl
from app import COACHES, PLAYERS, app
from tests.helpers import csrf_data


def login(client):
    return client.post(
        "/login",
        data={"username": "admin", "password": "admin123"},
        follow_redirects=True,
    )


def test_training_plan_excel_import_skips_records_already_exported():
    original_plans = deepcopy(app_module.TRAINING_PLANS)
    original_counter = app_module.PLAN_ID_COUNTER
    try:
        client = app.test_client()
        login(client)

        exported = client.get("/training/plans/export")
        payload = io.BytesIO(exported.data)

        response = client.post(
            "/training/import-excel",
            data=csrf_data(client, {"training_excel": (payload, "plans.xlsx")}),
            content_type="multipart/form-data",
        )

        assert response.status_code == 302
        assert len(app_module.TRAINING_PLANS) == len(original_plans)
    finally:
        app_module.TRAINING_PLANS[:] = original_plans
        app_module.PLAN_ID_COUNTER = original_counter


def test_training_plan_excel_import_skips_duplicate_rows_in_same_file():
    original_plans = deepcopy(app_module.TRAINING_PLANS)
    original_counter = app_module.PLAN_ID_COUNTER
    try:
        client = app.test_client()
        login(client)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["运动员", "教练", "训练日期", "内容", "强度", "时长(分钟)", "地点"])
        duplicate_row = [
            PLAYERS[0]["name"],
            COACHES[0]["name"],
            "2026-07-20 09:00",
            "同一文件重复计划",
            "中",
            60,
            "训练馆A",
        ]
        sheet.append(duplicate_row)
        sheet.append(duplicate_row)
        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)

        response = client.post(
            "/training/import-excel",
            data=csrf_data(client, {"training_excel": (payload, "plans.xlsx")} ),
            content_type="multipart/form-data",
            follow_redirects=True,
        )

        assert response.status_code == 200
        assert len(app_module.TRAINING_PLANS) == len(original_plans) + 1
        assert "第3行：重复训练计划已跳过" in response.get_data(as_text=True)
    finally:
        app_module.TRAINING_PLANS[:] = original_plans
        app_module.PLAN_ID_COUNTER = original_counter


def test_training_plan_excel_import_redirects_to_plan_list():
    original_plans = deepcopy(app_module.TRAINING_PLANS)
    original_counter = app_module.PLAN_ID_COUNTER
    try:
        client = app.test_client()
        login(client)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["运动员", "教练", "训练日期", "内容", "强度", "时长(分钟)", "地点"])
        sheet.append([
            PLAYERS[0]["name"],
            COACHES[0]["name"],
            "2026-07-20 09:00",
            "导入回归测试",
            "中",
            60,
            "训练馆A",
        ])
        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)

        response = client.post(
            "/training/import-excel",
            data=csrf_data(client, {"training_excel": (payload, "plans.xlsx")}),
            content_type="multipart/form-data",
            follow_redirects=False,
        )

        assert response.status_code == 302
        assert response.headers["Location"].endswith("/training/plans")
    finally:
        app_module.TRAINING_PLANS[:] = original_plans
        app_module.PLAN_ID_COUNTER = original_counter


def test_training_plan_excel_import_rejects_skill_record_template():
    original_plans = deepcopy(app_module.TRAINING_PLANS)
    original_counter = app_module.PLAN_ID_COUNTER
    try:
        client = app.test_client()
        login(client)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "专项技术记录"
        sheet.append(["运动员", "训练日期", "步法训练", "击球技术", "多球时长", "训练强度", "备注"])
        sheet.append([
            PLAYERS[0]["name"],
            "2026-07-20",
            "综合步法",
            "防守转换",
            48,
            "极高强度",
            "专项技术记录误传到训练计划入口",
        ])
        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)

        response = client.post(
            "/training/import-excel",
            data=csrf_data(client, {"training_excel": (payload, "skill_records.xlsx")}),
            content_type="multipart/form-data",
            follow_redirects=True,
        )

        body = response.get_data(as_text=True)
        assert response.status_code == 200
        assert len(app_module.TRAINING_PLANS) == len(original_plans)
        assert "该文件是步法训练记录模板" in body
        assert "30 个错误" not in body
    finally:
        app_module.TRAINING_PLANS[:] = original_plans
        app_module.PLAN_ID_COUNTER = original_counter
