import io
from copy import deepcopy

import openpyxl

import app as app_module
from app import FOOTWORK_TRAINING_RECORDS, app
from tests.helpers import csrf_data


def login(client, username="admin", password="admin123"):
    return client.post(
        "/login",
        data={"username": username, "password": password},
        follow_redirects=True,
    )


def test_stats_dashboard_renders_charts_and_cards():
    client = app.test_client()
    login(client)

    response = client.get("/stats/dashboard")
    body = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "数据统计看板" in body
    assert "月度训练时长趋势" in body
    assert "伤病部位分布" in body
    assert "trainTrendChart" in body
    for chart_id in (
        "fitnessDetailChart",
        "radarChart",
        "footworkTrendChart",
        "skillTrendChart",
        "intensityPieChart",
    ):
        assert chart_id in body


def test_build_overall_stats_provides_dashboard_schema():
    stats = app_module.build_overall_stats()

    for card_key in ("total_athletes", "current_month_duration", "active_injuries", "avg_fitness", "avg_tech", "avg_footwork"):
        assert card_key in stats["cards"]

    assert set(stats["fitness_detail"]) == {
        "names",
        "upper_strength",
        "lower_strength",
        "flexibility",
        "endurance",
        "speed",
    }
    assert set(stats["radar"]) == {"names", "data"}
    assert "footwork_month_labels" in stats
    assert "footwork_month_scores" in stats


def test_build_overall_stats_uses_training_repository_and_current_score_logic(monkeypatch):
    player = {"id": 101, "name": "测试运动员", "gender": "男"}
    fitness_record = app_module.build_fitness_test_record(
        501,
        player["id"],
        "2026-07-01 09:00",
        2,
        {
            "sprint_30m": 4.5,
            "abdominal_endurance": 100,
            "back_endurance": 100,
            "lateral_slide": 7.0,
            "a_footwork": 11.0,
            "double_under": 100,
            "seated_rotation_throw": 500,
            "standing_long_jump": 240,
        },
        "",
        "admin",
    )
    repository_footwork_records = [
        {
            "id": 1,
            "athlete_id": player["id"],
            "training_date": "2026-07-02",
            "footwork_dict_id": 1,
            "duration_minutes": 40,
            "set_count": 4,
            "note": "",
        }
    ]
    repository_technique_records = [
        {
            "id": 2,
            "athlete_id": player["id"],
            "training_date": "2026-07-03",
            "technique_dict_id": 1,
            "technique_category_id": 1,
            "multi_ball_count": 120,
            "serve_frequency": "高",
            "plan_execution_rate": 70,
            "on_table_rate": 80,
            "hit_score": 12,
            "landing_items": [],
            "qualitative_comment": "",
        }
    ]

    monkeypatch.setattr(app_module, "PLAYERS", [player])
    monkeypatch.setattr(app_module, "FITNESS_TESTS", [fitness_record])
    monkeypatch.setattr(app_module, "TRAINING_PLANS", [])
    monkeypatch.setattr(app_module, "INJURY_RECORDS", [])
    monkeypatch.setattr(
        app_module.training_repo,
        "list_footwork_records",
        lambda filters=None: repository_footwork_records,
    )
    monkeypatch.setattr(
        app_module.training_repo,
        "list_technique_tactic_records",
        lambda filters=None: repository_technique_records,
    )

    stats = app_module.build_overall_stats()

    assert stats["cards"]["avg_tech"] == 80
    assert stats["cards"]["avg_footwork"] == 100
    assert stats["skill_month_labels"] == ["2026-07"]
    assert stats["skill_month_scores"] == [80]
    assert stats["footwork_month_labels"] == ["2026-07"]
    assert stats["footwork_month_scores"] == [40]


def test_stats_export_all_returns_xlsx_workbook():
    client = app.test_client()
    login(client)

    response = client.get("/stats/export/all")

    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = openpyxl.load_workbook(io.BytesIO(response.data))
    assert workbook.sheetnames == ["训练计划", "步法训练记录", "技战术训练记录", "体能训练记录", "伤病记录"]


def test_stats_import_accepts_member9_skill_excel_format():
    client = app.test_client()
    login(client)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["运动员姓名", "训练日期", "步法时长", "击球得分", "多球时长", "训练强度"])
    sheet.append(["王一鸣", "2026-07-09", 45, 88, 30, "高"])
    payload = io.BytesIO()
    workbook.save(payload)
    payload.seek(0)

    before_count = len(FOOTWORK_TRAINING_RECORDS)
    try:
        response = client.post(
            "/stats/import-export",
            data=csrf_data(
                client,
                {
                    "action": "import_skill",
                    "skill_excel": (payload, "member9_skill.xlsx"),
                },
            ),
            content_type="multipart/form-data",
            follow_redirects=True,
        )

        assert response.status_code == 200
        assert len(FOOTWORK_TRAINING_RECORDS) == before_count + 1
        imported = FOOTWORK_TRAINING_RECORDS[-1]
        assert imported["athlete_name"] == "王一鸣"
        assert imported["training_date"] == "2026-07-09"
        assert imported["duration_minutes"] == 45
    finally:
        del FOOTWORK_TRAINING_RECORDS[before_count:]


def test_stats_import_creates_player_archive_for_new_names():
    client = app.test_client()
    login(client)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["运动员姓名", "训练日期", "步法时长", "击球得分", "多球时长", "训练强度"])
    sheet.append(["小红", "2026-06-09", 27, 82, 27, "低"])
    payload = io.BytesIO()
    workbook.save(payload)
    payload.seek(0)

    original_players = deepcopy(app_module.PLAYERS)
    before_count = len(FOOTWORK_TRAINING_RECORDS)
    try:
        response = client.post(
            "/stats/import-export",
            data=csrf_data(
                client,
                {
                    "action": "import_skill",
                    "skill_excel": (payload, "member9_skill_new_player.xlsx"),
                },
            ),
            content_type="multipart/form-data",
            follow_redirects=True,
        )

        created_player = next((player for player in app_module.PLAYERS if player["name"] == "小红"), None)
        imported = FOOTWORK_TRAINING_RECORDS[-1]

        assert response.status_code == 200
        assert created_player is not None
        assert created_player["student_no"].startswith("IMP")
        assert len(FOOTWORK_TRAINING_RECORDS) == before_count + 1
        assert imported["athlete_id"] == created_player["id"]
        assert imported["athlete_name"] == "小红"
    finally:
        app_module.PLAYERS[:] = original_players
        del FOOTWORK_TRAINING_RECORDS[before_count:]
