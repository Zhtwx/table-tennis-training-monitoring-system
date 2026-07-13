import io
import os
import secrets
import sys
import openpyxl
from flask import send_file
from openpyxl.styles import Font, Alignment
from copy import deepcopy
from datetime import datetime, timedelta

from flask import (
    Blueprint,
    Flask,
    flash,
    redirect,
    render_template,
    request,
    session,
    url_for,
)

from auth_utils import (
    USERS,
    can_delete_training_plan,
    current_user,
    is_safe_redirect_url,
    role_required,
)
from repositories import training_repository as training_repo
from security import AUDIT_LOGS, csrf_token, record_audit_log, validate_csrf_token

NAV_GROUPS = [
    {
        "label": "训练业务",
        "items": [
            {"label": "比赛成绩", "endpoint": "matches.list", "roles": {"admin", "coach"}},
            {"label": "训练计划", "endpoint": "training.plans", "roles": {"admin", "coach"}},
            {"label": "体能训练", "endpoint": "fitness.training", "roles": {"admin", "coach"}},
            {"label": "步法训练", "endpoint": "training.footwork", "roles": {"admin", "coach"}},
            {"label": "技战术训练", "endpoint": "training.technique_tactic", "roles": {"admin", "coach"}},
        ],
    },
    {
        "label": "人员与状态",
        "items": [
            {"label": "运动员档案", "endpoint": "players.list", "roles": {"admin", "coach"}},
            {"label": "教练员信息", "endpoint": "coaches.list", "roles": {"admin", "coach"}},
            {
                "label": "运动员身体状态",
                "children": [
                    {"label": "伤病记录", "endpoint": "injuries.list", "roles": {"admin", "coach"}},
                    {"label": "康复跟踪", "endpoint": "rehab.list", "roles": {"admin", "coach"}},
                ],
            },
        ],
    },
    {
        "label": "数据与管理",
        "items": [
            {"label": "数据统计", "endpoint": "stats.dashboard", "roles": {"admin", "coach"}},
            {"label": "导入导出", "endpoint": "stats.import_export", "roles": {"admin", "coach"}},
            {"label": "用户权限", "endpoint": "auth.users", "roles": {"admin"}},
        ],
    },
]


def build_navigation_for_role(role):
    nav_groups = []
    nav_items = []
    for group in NAV_GROUPS:
        visible_items = []
        for item in group["items"]:
            item_roles = item.get("roles", set())
            if role not in item_roles and "children" not in item:
                continue

            visible_item = {key: value for key, value in item.items() if key != "children"}
            children = item.get("children")
            if children:
                visible_children = [
                    {key: value for key, value in child.items()}
                    for child in children
                    if role in child.get("roles", set())
                ]
                if not visible_children:
                    continue
                visible_item["children"] = visible_children
                nav_items.extend(visible_children)
            else:
                nav_items.append(visible_item)
            visible_items.append(visible_item)

        if visible_items:
            nav_groups.append({"label": group["label"], "items": visible_items})

    return nav_groups, nav_items

PLAYERS = [
    {
        "id": 1,
        "student_no": "2026001",
        "name": "王一鸣",
        "gender": "男",
        "age": 19,
        "level": "一级运动员",
        "skill_level": "一级运动员",
        "level_code": "first",
        "play_style": "右手横板快攻结合弧圈",
        "injury_status": "健康",
        "injury_status_code": "healthy",
    },
    {
        "id": 2,
        "student_no": "2026002",
        "name": "李清扬",
        "gender": "女",
        "age": 18,
        "level": "二级运动员",
        "skill_level": "二级运动员",
        "level_code": "second",
        "play_style": "左手横板两面弧圈",
        "injury_status": "观察中",
        "injury_status_code": "observe",
    },
    {
        "id": 3,
        "student_no": "2026003",
        "name": "陈昊然",
        "gender": "男",
        "age": 21,
        "level": "国家级",
        "skill_level": "国家级",
        "level_code": "national",
        "play_style": "右手直板近台快攻",
        "injury_status": "康复中",
        "injury_status_code": "rehab",
    },
    {
        "id": 4,
        "student_no": "2026004",
        "name": "赵若溪",
        "gender": "女",
        "age": 20,
        "level": "一级运动员",
        "skill_level": "一级运动员",
        "level_code": "first",
        "play_style": "右手横板反手快拨",
        "injury_status": "伤病中",
        "injury_status_code": "injured",
    },
]
# 教练模拟数据
COACHES = [
    {"id": 1, "name": "张教练", "gender": "男", "specialty": "乒乓球专项训练"},
    {"id": 2, "name": "李教练", "gender": "女", "specialty": "体能训练与康复"},
]

# 训练计划模拟数据（内存存储）
TRAINING_PLANS = [
    {
        "id": 1,
        "athlete_id": 1,
        "athlete_name": "王一鸣",
        "coach_id": 1,
        "coach_name": "张教练",
        "plan_datetime": "2026-07-01 09:00",
        "content": "正手攻球+反手推挡",
        "intensity": "高",
        "duration": 90,
        "location": "训练馆A",
    },
    {
        "id": 2,
        "athlete_id": 2,
        "athlete_name": "李清扬",
        "coach_id": 2,
        "coach_name": "李教练",
        "plan_datetime": "2026-07-02 10:00",
        "content": "发球抢攻战术",
        "intensity": "中",
        "duration": 75,
        "location": "训练馆B",
    },
]
PLAN_ID_COUNTER = 3
TRAINING_PLAN_ITEMS = []
PLAN_ITEM_ID_COUNTER = 1
TRAINING_PLAN_STATUS_LABELS = {
    "draft": "草稿",
    "published": "已发布",
    "running": "执行中",
    "completed": "已完成",
    "cancelled": "已取消",
}
EXECUTABLE_PLAN_STATUSES = {"published", "running", "completed"}

FOOTWORK_TRAINING_RECORDS = training_repo.MEMORY_FOOTWORK_RECORDS
TECHNIQUE_TACTIC_TRAINING_RECORDS = training_repo.MEMORY_TECHNIQUE_TACTIC_RECORDS
SERVE_FREQUENCY_OPTIONS = training_repo.SERVE_FREQUENCY_OPTIONS
LANDING_CONCENTRATION_OPTIONS = training_repo.LANDING_CONCENTRATION_OPTIONS
LANDING_DISTRIBUTION_OPTIONS = training_repo.LANDING_DISTRIBUTION_OPTIONS
SYS_DICTIONARY = training_repo.SYS_DICTIONARY

PLAYER_LEVEL_LABELS = {
    "national": "国家级",
    "first": "一级运动员",
    "second": "二级运动员",
    "youth": "青年队",
}

PLAYER_INJURY_STATUS_LABELS = {
    "healthy": "健康",
    "observe": "观察中",
    "rehab": "康复中",
    "injured": "伤病中",
}

PLAYER_GENDER_OPTIONS = ["男", "女"]



FITNESS_TESTS = [
    {
        "id": 1,
        "athlete_id": 1,
        "test_date": "2026-06-03 9:30",
        "tester_id": 2,
        "upper_strength": 84.0,
        "lower_strength": 88.0,
        "flexibility": 82.0,
        "endurance": 86.0,
        "speed": 91.0,
        "overall_score": 86.2,
        "notes": "训练状态稳定，体能结构均衡。",
        "created_by": "coach",
    },
    {
        "id": 2,
        "athlete_id": 2,
        "test_date": "2026-06-11 14:00",
        "tester_id": 2,
        "upper_strength": 76.0,
        "lower_strength": 78.0,
        "flexibility": 74.0,
        "endurance": 79.0,
        "speed": 73.0,
        "overall_score": 76.0,
        "notes": "速度指标偏低，建议增加敏捷与启动练习。",
        "created_by": "coach",
    },
    {
        "id": 3,
        "athlete_id": 3,
        "test_date": "2026-06-18 10:00",
        "tester_id": 2,
        "upper_strength": 82.0,
        "lower_strength": 84.0,
        "flexibility": 68.0,
        "endurance": 81.0,
        "speed": 76.0,
        "overall_score": 78.2,
        "notes": "柔韧性偏低，需加强拉伸和恢复。",
        "created_by": "admin",
    },
    {
        "id": 4,
        "athlete_id": 4,
        "test_date": "2026-07-02 15:00",
        "tester_id": 2,
        "upper_strength": 70.0,
        "lower_strength": 66.0,
        "flexibility": 58.0,
        "endurance": 69.0,
        "speed": 64.0,
        "overall_score": 65.4,
        "notes": "恢复期指标偏弱，维持低强度过渡方案。",
        "created_by": "coach",
    },
]

TRAINING_SYNC_LOGS = [
    {
        "id": 1,
        "fitness_test_id": 1,
        "athlete_id": 1,
        "coach_id": 2,
        "sync_date": "2026-06-03 ",
        "plan_name": "体能巩固训练",
        "hours": 16.0,
        "intensity": "中",
        "status": "已完成",
    },
    {
        "id": 2,
        "fitness_test_id": 2,
        "athlete_id": 2,
        "coach_id": 2,
        "sync_date": "2026-06-11",
        "plan_name": "速度敏捷提升",
        "hours": 18.0,
        "intensity": "中",
        "status": "已完成",
    },
    {
        "id": 3,
        "fitness_test_id": 3,
        "athlete_id": 3,
        "coach_id": 2,
        "sync_date": "2026-06-18",
        "plan_name": "恢复拉伸结合耐力课",
        "hours": 20.0,
        "intensity": "高",
        "status": "进行中",
    },
    {
        "id": 4,
        "fitness_test_id": 4,
        "athlete_id": 4,
        "coach_id": 2,
        "sync_date": "2026-07-02",
        "plan_name": "康复过渡训练",
        "hours": 12.0,
        "intensity": "低",
        "status": "进行中",
    },
]

INJURY_RECORDS = [
    {
        "id": 1,
        "athlete_id": 1,
        "injury_date": "2026-05-10",
        "injury_location": "右肩",
        "injury_type": "肌肉拉伤",
        "severity": "中度",
        "diagnosis": "肩袖肌腱炎，建议理疗并减少高强度上肢训练。",
        "treatment": "超声波理疗，弹力带康复训练。",
        "recovery_status": "已恢复",
        "expected_recovery_date": "2026-05-28",
        "notes": "恢复后两周内控制发力训练量。",
        "created_by": "coach",
        "is_deleted": False,
        "deleted_by": "",
        "deleted_at": "",
        "delete_reason": "",
    },
    {
        "id": 2,
        "athlete_id": 2,
        "injury_date": "2026-06-22",
        "injury_location": "右腕",
        "injury_type": "过度使用疼痛",
        "severity": "轻微",
        "diagnosis": "腕部负荷偏高，击球后疼痛明显。",
        "treatment": "冰敷、护腕保护，降低反手连续训练量。",
        "recovery_status": "治疗中",
        "expected_recovery_date": "2026-07-12",
        "notes": "继续观察握拍稳定性。",
        "created_by": "coach",
        "is_deleted": False,
        "deleted_by": "",
        "deleted_at": "",
        "delete_reason": "",
    },
    {
        "id": 3,
        "athlete_id": 3,
        "injury_date": "2026-06-30",
        "injury_location": "腰椎",
        "injury_type": "椎间盘轻微突出",
        "severity": "严重",
        "diagnosis": "L4-L5椎间盘轻微突出，已停止高冲击训练。",
        "treatment": "牵引、核心肌群康复训练，配合低强度步法恢复。",
        "recovery_status": "康复中",
        "expected_recovery_date": "2026-07-30",
        "notes": "复训前需完成专项评估。",
        "created_by": "admin",
        "is_deleted": False,
        "deleted_by": "",
        "deleted_at": "",
        "delete_reason": "",
    },
    {
        "id": 4,
        "athlete_id": 4,
        "injury_date": "2026-06-15",
        "injury_location": "左膝",
        "injury_type": "韧带扭伤",
        "severity": "严重",
        "diagnosis": "左膝内侧副韧带二级扭伤，短期不适合对抗训练。",
        "treatment": "制动休息、消肿处理、康复师指导活动度训练。",
        "recovery_status": "治疗中",
        "expected_recovery_date": "2026-07-20",
        "notes": "禁止多球大范围移动训练。",
        "created_by": "coach",
        "is_deleted": False,
        "deleted_by": "",
        "deleted_at": "",
        "delete_reason": "",
    },
]

INJURY_FOLLOWUPS = [
    {
        "id": 1,
        "injury_record_id": 2,
        "followup_date": "2026-07-02",
        "pain_score": 3,
        "training_limit": "避免连续反手发力，单次训练不超过 60 分钟。",
        "advice": "继续冰敷和护腕保护，三天后复查握拍疼痛。",
        "reviewer": "刘指导",
        "created_by": "coach",
    },
    {
        "id": 2,
        "injury_record_id": 3,
        "followup_date": "2026-07-05",
        "pain_score": 4,
        "training_limit": "禁止高冲击步法和大幅度侧身进攻。",
        "advice": "保持核心稳定训练，复训前完成腰椎活动度评估。",
        "reviewer": "陈指导",
        "created_by": "admin",
    },
]

INJURY_SEVERITY_OPTIONS = ["轻微", "中度", "严重"]
INJURY_RECOVERY_STATUS_OPTIONS = ["治疗中", "康复中", "已恢复"]

ATHLETE_INJURY_STATUS_META = {
    "健康": {"code": "healthy", "class": "success"},
    "观察中": {"code": "observe", "class": "warning"},
    "康复中": {"code": "rehab", "class": "info"},
    "伤病中": {"code": "injured", "class": "danger"},
}

INTENSITY_LABELS = {
    "低": "低",
    "中": "中",
    "高": "高",
    "极高": "极高",
}

MATCH_RESULTS = [
    {
        "id": 1,
        "athlete_id": 1,
        "match_date": "2026-05-15",
        "match_name": "2026全国大学生锦标赛-男单",
        "opponent": "刘强",
        "result": "胜",
        "score": "3:1",
        "rank": "八强",
        "key_points": "关键分 7:4，发球轮次连续得分。",
        "tactic_review": "正手位拉球质量高，前三板衔接稳定。",
        "improvement": "反手防守落点需要更主动。",
    },
    {
        "id": 2,
        "athlete_id": 1,
        "match_date": "2026-05-17",
        "match_name": "2026全国大学生锦标赛-男单",
        "opponent": "何伟",
        "result": "负",
        "score": "1:3",
        "rank": "八强",
        "key_points": "第二局 9:9 后连续丢接发球分。",
        "tactic_review": "对手反手快撕压制明显，台内球处理偏保守。",
        "improvement": "补强反手位防守转换和接发抢攻线路。",
    },
    {
        "id": 3,
        "athlete_id": 2,
        "match_date": "2026-06-02",
        "match_name": "省青联赛第一轮",
        "opponent": "周敏",
        "result": "胜",
        "score": "3:2",
        "rank": "小组第2",
        "key_points": "决胜局 6:8 反超，连续三板相持得分。",
        "tactic_review": "左手线路变化有效，反手拧拉打开局面。",
        "improvement": "领先后需要减少无谓抢攻失误。",
    },
    {
        "id": 4,
        "athlete_id": 3,
        "match_date": "2026-06-15",
        "match_name": "省青联赛第一轮",
        "opponent": "李浩",
        "result": "负",
        "score": "2:3",
        "rank": "小组第3",
        "key_points": "决胜局 8:10 追回一分后被反拉得分。",
        "tactic_review": "正手主动上手足够，但腰椎伤情影响侧身连续性。",
        "improvement": "结合康复进度控制侧身进攻比例。",
    },
    {
        "id": 5,
        "athlete_id": 4,
        "match_date": "2026-06-20",
        "match_name": "校际交流赛",
        "opponent": "林悦",
        "result": "胜",
        "score": "3:0",
        "rank": "团体赛出场",
        "key_points": "发抢轮次得分率高，第三局连续拿下 5 分。",
        "tactic_review": "反手快拨衔接清晰，线路变化充分。",
        "improvement": "膝伤恢复期不安排长回合对拉。",
    },
]
MATCH_TACTICAL_ANALYSES = []
MATCH_PHASE_STATS = []
MATCH_ANALYSIS_ID_COUNTER = 1
MATCH_PHASE_STAT_ID_COUNTER = 1

DEMO_PLAYERS = [
    {
        "id": 5,
        "student_no": "2026005",
        "name": "孙泽宇",
        "gender": "男",
        "age": 17,
        "level": "青年队",
        "skill_level": "青年队",
        "level_code": "youth",
        "play_style": "右手横板快攻弧圈",
        "injury_status": "健康",
        "injury_status_code": "healthy",
    },
    {
        "id": 6,
        "student_no": "2026006",
        "name": "周雨桐",
        "gender": "女",
        "age": 19,
        "level": "一级运动员",
        "skill_level": "一级运动员",
        "level_code": "first",
        "play_style": "右手削攻结合",
        "injury_status": "观察中",
        "injury_status_code": "observe",
    },
    {
        "id": 7,
        "student_no": "2026007",
        "name": "吴嘉宁",
        "gender": "男",
        "age": 20,
        "level": "二级运动员",
        "skill_level": "二级运动员",
        "level_code": "second",
        "play_style": "左手直板近台快攻",
        "injury_status": "健康",
        "injury_status_code": "healthy",
    },
    {
        "id": 8,
        "student_no": "2026008",
        "name": "郑可欣",
        "gender": "女",
        "age": 18,
        "level": "青年队",
        "skill_level": "青年队",
        "level_code": "youth",
        "play_style": "右手横板两面弧圈",
        "injury_status": "康复中",
        "injury_status_code": "rehab",
    },
]

DEMO_TRAINING_PLANS = [
    {
        "id": 3,
        "athlete_id": 5,
        "athlete_name": "孙泽宇",
        "coach_id": 1,
        "coach_name": "张教练",
        "plan_datetime": "2026-07-03 09:30",
        "content": "正手连续弧圈与落点控制",
        "intensity": "高",
        "duration": 85,
        "location": "训练馆A",
    },
    {
        "id": 4,
        "athlete_id": 6,
        "athlete_name": "周雨桐",
        "coach_id": 2,
        "coach_name": "李教练",
        "plan_datetime": "2026-07-04 14:00",
        "content": "削球稳定性与防守反击转换",
        "intensity": "中",
        "duration": 70,
        "location": "训练馆B",
    },
    {
        "id": 5,
        "athlete_id": 7,
        "athlete_name": "吴嘉宁",
        "coach_id": 1,
        "coach_name": "张教练",
        "plan_datetime": "2026-07-06 10:00",
        "content": "直板前三板抢攻与台内挑打",
        "intensity": "高",
        "duration": 95,
        "location": "训练馆A",
    },
    {
        "id": 6,
        "athlete_id": 8,
        "athlete_name": "郑可欣",
        "coach_id": 2,
        "coach_name": "李教练",
        "plan_datetime": "2026-07-08 16:00",
        "content": "康复期步法调整与低负荷多球",
        "intensity": "低",
        "duration": 60,
        "location": "康复训练区",
    },
    {
        "id": 7,
        "athlete_id": 5,
        "athlete_name": "孙泽宇",
        "coach_id": 1,
        "coach_name": "张教练",
        "plan_datetime": "2026-07-10 09:00",
        "content": "发球抢攻与反手衔接强化",
        "intensity": "极高",
        "duration": 105,
        "location": "训练馆C",
    },
    {
        "id": 8,
        "athlete_id": 7,
        "athlete_name": "吴嘉宁",
        "coach_id": 2,
        "coach_name": "李教练",
        "plan_datetime": "2026-07-12 15:30",
        "content": "赛前速度耐力与连续变线",
        "intensity": "中",
        "duration": 65,
        "location": "训练馆B",
    },
]

DEMO_FITNESS_TESTS = [
    {
        "id": 5,
        "athlete_id": 5,
        "test_date": "2026-07-03 11:00",
        "tester_id": 2,
        "upper_strength": 82.0,
        "lower_strength": 86.0,
        "flexibility": 78.0,
        "endurance": 84.0,
        "speed": 88.0,
        "overall_score": 83.6,
        "notes": "速度和下肢力量突出，柔韧性仍有提升空间。",
        "created_by": "coach",
    },
    {
        "id": 6,
        "athlete_id": 6,
        "test_date": "2026-07-04 15:30",
        "tester_id": 2,
        "upper_strength": 74.0,
        "lower_strength": 77.0,
        "flexibility": 81.0,
        "endurance": 80.0,
        "speed": 72.0,
        "overall_score": 76.8,
        "notes": "速度偏弱，结合腕部观察期控制上肢负荷。",
        "created_by": "coach",
    },
    {
        "id": 7,
        "athlete_id": 7,
        "test_date": "2026-07-12 17:00",
        "tester_id": 2,
        "upper_strength": 79.0,
        "lower_strength": 83.0,
        "flexibility": 76.0,
        "endurance": 82.0,
        "speed": 85.0,
        "overall_score": 81.0,
        "notes": "整体稳定，适合进入赛前专项速度耐力强化。",
        "created_by": "admin",
    },
]

DEMO_INJURY_RECORDS = [
    {
        "id": 5,
        "athlete_id": 6,
        "injury_date": "2026-07-04",
        "injury_location": "左腕",
        "injury_type": "削球负荷性疼痛",
        "severity": "轻微",
        "diagnosis": "连续削球训练后腕背疼痛，未见明显肿胀。",
        "treatment": "降低削球连续组数，训练后冰敷并佩戴护腕。",
        "recovery_status": "治疗中",
        "expected_recovery_date": "2026-07-18",
        "notes": "观察握拍发力角度，避免连续高负荷反手削球。",
        "created_by": "coach",
        "is_deleted": False,
        "deleted_by": "",
        "deleted_at": "",
        "delete_reason": "",
    },
    {
        "id": 6,
        "athlete_id": 8,
        "injury_date": "2026-07-01",
        "injury_location": "右踝",
        "injury_type": "轻度扭伤恢复期",
        "severity": "中度",
        "diagnosis": "右踝外侧韧带轻度扭伤后恢复期，跳步仍有不适。",
        "treatment": "踝关节稳定训练，限制大范围交叉步移动。",
        "recovery_status": "康复中",
        "expected_recovery_date": "2026-07-25",
        "notes": "低强度多球配合康复评估，逐步恢复对抗。",
        "created_by": "coach",
        "is_deleted": False,
        "deleted_by": "",
        "deleted_at": "",
        "delete_reason": "",
    },
]

DEMO_MATCH_RESULTS = [
    {
        "id": 6,
        "athlete_id": 5,
        "match_date": "2026-07-11",
        "match_name": "市青少年积分赛",
        "opponent": "郑远",
        "result": "胜",
        "score": "3:1",
        "rank": "四强",
        "key_points": "第三局 8:8 后通过发球抢攻连续拿下两分。",
        "tactic_review": "正手弧圈落点变化充分，反手衔接比训练初期更稳定。",
        "improvement": "领先阶段需要减少冒险接发抢攻。",
    },
]

ADDITIONAL_ATHLETE_PROFILES = [
    {"name": "韩子昂", "gender": "男", "age": 18, "level_code": "youth", "play_style": "右手横板正手抢攻"},
    {"name": "林思琪", "gender": "女", "age": 19, "level_code": "second", "play_style": "左手横板两面弧圈"},
    {"name": "许明远", "gender": "男", "age": 20, "level_code": "first", "play_style": "右手直板快攻结合推挡"},
    {"name": "罗佳怡", "gender": "女", "age": 17, "level_code": "youth", "play_style": "右手横板防守反击"},
    {"name": "高俊熙", "gender": "男", "age": 21, "level_code": "first", "play_style": "右手横板中远台弧圈"},
    {"name": "唐婉宁", "gender": "女", "age": 18, "level_code": "second", "play_style": "右手横板快带快撕"},
    {"name": "魏晨皓", "gender": "男", "age": 19, "level_code": "youth", "play_style": "左手直板前三板抢攻"},
    {"name": "沈若彤", "gender": "女", "age": 20, "level_code": "first", "play_style": "右手横板削攻结合"},
    {"name": "蒋宇航", "gender": "男", "age": 18, "level_code": "second", "play_style": "右手横板反手拧拉"},
    {"name": "马依然", "gender": "女", "age": 21, "level_code": "first", "play_style": "左手横板落点变化"},
    {"name": "邓凯文", "gender": "男", "age": 17, "level_code": "youth", "play_style": "右手直板近台快攻"},
    {"name": "方语桐", "gender": "女", "age": 19, "level_code": "second", "play_style": "右手横板发抢衔接"},
    {"name": "秦浩宇", "gender": "男", "age": 20, "level_code": "first", "play_style": "右手横板力量弧圈"},
    {"name": "白欣妍", "gender": "女", "age": 18, "level_code": "youth", "play_style": "左手横板快攻快带"},
    {"name": "丁睿泽", "gender": "男", "age": 21, "level_code": "first", "play_style": "右手横板相持控制"},
    {"name": "姚可薇", "gender": "女", "age": 20, "level_code": "second", "play_style": "右手横板反手压制"},
    {"name": "孟泽霖", "gender": "男", "age": 18, "level_code": "youth", "play_style": "左手横板发球变化"},
    {"name": "程安琪", "gender": "女", "age": 19, "level_code": "first", "play_style": "右手横板两面快拉"},
    {"name": "薛景辰", "gender": "男", "age": 20, "level_code": "second", "play_style": "右手直板台内控制"},
    {"name": "梁沐瑶", "gender": "女", "age": 17, "level_code": "youth", "play_style": "右手横板速度型打法"},
    {"name": "叶承佑", "gender": "男", "age": 21, "level_code": "first", "play_style": "左手横板正反手均衡"},
    {"name": "苏芷晴", "gender": "女", "age": 18, "level_code": "second", "play_style": "右手横板防守转换"},
]
ADDITIONAL_ATHLETE_COUNT = len(ADDITIONAL_ATHLETE_PROFILES)

_ADDITIONAL_LEVEL_NAMES = {
    "national": "国家级",
    "first": "一级运动员",
    "second": "二级运动员",
    "youth": "青年队",
}
_ADDITIONAL_STATUS_SEQUENCE = [
    ("健康", "healthy"),
    ("观察中", "observe"),
    ("健康", "healthy"),
    ("康复中", "rehab"),
]
_ADDITIONAL_INTENSITIES = ["低", "中", "高", "极高"]
_ADDITIONAL_TECHNICAL_INTENSITIES = ["low", "medium", "high", "extreme"]
_ADDITIONAL_TREND_MONTHS = [
    "2025-10",
    "2025-11",
    "2025-12",
    "2026-01",
    "2026-02",
    "2026-03",
    "2026-04",
    "2026-05",
    "2026-06",
    "2026-07",
]
_ADDITIONAL_FOOTWORK_CODES = ["single_step", "parallel_step", "cross_step", "composite_step", "shuffle_step", "stride_step", "side_step", "recovery_step"]
_ADDITIONAL_TECHNIQUE_CODES = ["forehand_loop", "backhand_flick", "serve_attack", "receive_push_short", "change_direction", "defense_to_attack", "wide_angle", "block"]
_ADDITIONAL_INJURY_DETAILS = [
    (9, "右肩", "肩袖疲劳", "轻微", "治疗中"),
    (12, "左膝", "跳步落地不适", "中度", "康复中"),
    (15, "腰背", "核心疲劳性疼痛", "轻微", "治疗中"),
    (18, "右腕", "连续拧拉负荷反应", "轻微", "治疗中"),
    (21, "左踝", "移动急停扭伤", "中度", "康复中"),
    (24, "右肘", "发球训练后酸痛", "轻微", "治疗中"),
    (27, "大腿后侧", "肌肉拉伤恢复期", "中度", "康复中"),
    (30, "左肩", "防守转换拉伤", "轻微", "治疗中"),
]
_ADDITIONAL_MATCH_OPPONENTS = [
    "陈越",
    "赵宁",
    "顾然",
    "陆佳",
    "田峰",
    "何曼",
    "熊磊",
    "曹悦",
    "石博",
    "刘宸",
    "冯雪",
]

ADDITIONAL_DEMO_PLAYERS = [
    {
        "id": index + 9,
        "student_no": f"2026{index + 9:03d}",
        "name": profile["name"],
        "gender": profile["gender"],
        "age": profile["age"],
        "level": _ADDITIONAL_LEVEL_NAMES[profile["level_code"]],
        "skill_level": _ADDITIONAL_LEVEL_NAMES[profile["level_code"]],
        "level_code": profile["level_code"],
        "play_style": profile["play_style"],
        "injury_status": _ADDITIONAL_STATUS_SEQUENCE[index % len(_ADDITIONAL_STATUS_SEQUENCE)][0],
        "injury_status_code": _ADDITIONAL_STATUS_SEQUENCE[index % len(_ADDITIONAL_STATUS_SEQUENCE)][1],
    }
    for index, profile in enumerate(ADDITIONAL_ATHLETE_PROFILES)
]

ADDITIONAL_DEMO_TRAINING_PLANS = [
    {
        "id": index + 9,
        "athlete_id": player["id"],
        "athlete_name": player["name"],
        "coach_id": 1 if index % 2 == 0 else 2,
        "coach_name": "张教练" if index % 2 == 0 else "李教练",
        "plan_datetime": (
            f"{_ADDITIONAL_TREND_MONTHS[index % len(_ADDITIONAL_TREND_MONTHS)]}"
            f"-{(index % 20) + 1:02d} {8 + (index % 8):02d}:30"
        ),
        "content": f"{player['play_style']}专项巩固与多球衔接",
        "intensity": _ADDITIONAL_INTENSITIES[index % len(_ADDITIONAL_INTENSITIES)],
        "duration": 65 + (index % 6) * 10,
        "location": f"训练馆{chr(ord('A') + index % 3)}",
    }
    for index, player in enumerate(ADDITIONAL_DEMO_PLAYERS)
]

ADDITIONAL_DEMO_FITNESS_TESTS = [
    {
        "id": index + 8,
        "athlete_id": player["id"],
        "test_date": (
            f"{_ADDITIONAL_TREND_MONTHS[index % len(_ADDITIONAL_TREND_MONTHS)]}"
            f"-{(index % 20) + 1:02d} {10 + (index % 6):02d}:00"
        ),
        "tester_id": 2,
        "upper_strength": float(72 + index % 15),
        "lower_strength": float(74 + (index * 2) % 15),
        "flexibility": float(70 + (index * 3) % 16),
        "endurance": float(76 + (index * 4) % 14),
        "speed": float(73 + (index * 5) % 17),
        "overall_score": round(
            (
                (72 + index % 15)
                + (74 + (index * 2) % 15)
                + (70 + (index * 3) % 16)
                + (76 + (index * 4) % 14)
                + (73 + (index * 5) % 17)
            )
            / 5,
            1,
        ),
        "notes": f"{player['name']}体能测试完成，后续按短板指标同步训练负荷。",
        "created_by": "coach" if index % 2 else "admin",
    }
    for index, player in enumerate(ADDITIONAL_DEMO_PLAYERS)
]

ADDITIONAL_DEMO_INJURY_RECORDS = [
    {
        "id": index + 7,
        "athlete_id": athlete_id,
        "injury_date": f"2026-07-{index + 3:02d}",
        "injury_location": location,
        "injury_type": injury_type,
        "severity": severity,
        "diagnosis": f"{location}{injury_type}，需结合训练负荷进行观察。",
        "treatment": "降低相关专项强度，训练后冰敷或康复拉伸。",
        "recovery_status": recovery_status,
        "expected_recovery_date": f"2026-07-{index + 18:02d}",
        "notes": "与训练计划联动控制负荷，复训前完成教练评估。",
        "created_by": "coach",
        "is_deleted": False,
        "deleted_by": "",
        "deleted_at": "",
        "delete_reason": "",
    }
    for index, (athlete_id, location, injury_type, severity, recovery_status) in enumerate(
        _ADDITIONAL_INJURY_DETAILS
    )
]

ADDITIONAL_DEMO_MATCH_RESULTS = [
    {
        "id": index + 7,
        "athlete_id": player["id"],
        "match_date": f"2026-07-{index + 2:02d}",
        "match_name": "队内积分循环赛",
        "opponent": _ADDITIONAL_MATCH_OPPONENTS[index],
        "result": ["胜", "负", "胜", "平"][index % 4],
        "score": ["3:1", "2:3", "3:0", "2:2"][index % 4],
        "rank": f"第{index + 1}轮",
        "key_points": "关键分处理和发接发轮次作为复盘重点。",
        "tactic_review": f"{player['name']}在比赛中执行了既定线路和节奏变化。",
        "improvement": "继续提高领先阶段稳定性和相持球落点质量。",
    }
    for index, player in enumerate(ADDITIONAL_DEMO_PLAYERS[:11])
]

DEMO_EXTENSION_RECORD_COUNT = (
    len(DEMO_PLAYERS)
    + len(DEMO_TRAINING_PLANS)
    + 3
    + 3
    + len(DEMO_FITNESS_TESTS)
    + len(DEMO_INJURY_RECORDS)
    + len(DEMO_MATCH_RESULTS)
)
ADDITIONAL_DEMO_RECORD_COUNT = (
    len(ADDITIONAL_DEMO_PLAYERS)
    + len(ADDITIONAL_DEMO_TRAINING_PLANS)
    + len(ADDITIONAL_DEMO_PLAYERS) // 2
    + len(ADDITIONAL_DEMO_PLAYERS) - len(ADDITIONAL_DEMO_PLAYERS) // 2
    + len(ADDITIONAL_DEMO_FITNESS_TESTS)
    + len(ADDITIONAL_DEMO_INJURY_RECORDS)
    + len(ADDITIONAL_DEMO_MATCH_RESULTS)
)

PLAYERS.extend(DEMO_PLAYERS)
PLAYERS.extend(ADDITIONAL_DEMO_PLAYERS)


def assign_primary_coaches():
    """Populate the required coach affiliation for all in-memory athlete records."""
    for index, player in enumerate(PLAYERS):
        coach = COACHES[index % len(COACHES)]
        player["coach_id"] = coach["id"]
        player["coach_name"] = coach["name"]


def find_dict_by_id(dict_id):
    return training_repo.find_dict_by_id(dict_id)


def find_dict_by_code(dict_code):
    return training_repo.find_dict_by_code(dict_code)


def get_dict_children(category_type, parent_id=0, leaf_only=False):
    return training_repo.get_dict_children(category_type, parent_id, leaf_only)


def get_footwork_options():
    return training_repo.get_footwork_options()


def get_technique_categories():
    return training_repo.get_technique_categories()


def get_technique_options():
    return training_repo.get_technique_options()


def get_landing_options():
    return training_repo.get_landing_options()


def resolve_dict_by_name(name, category_type, parent_id=None):
    return training_repo.resolve_dict_by_name(name, category_type, parent_id)


def refresh_dictionary_groups():
    DICTIONARY_GROUPS[2]["items"] = [item["dict_name"] for item in get_footwork_options()]
    DICTIONARY_GROUPS[3]["items"] = [
        f"{category['dict_name']}/{child['dict_name']}"
        for category in get_technique_categories()
        for child in get_dict_children("technique_tactic", category["id"])
    ]
    DICTIONARY_GROUPS[4]["items"] = [item["dict_name"] for item in get_landing_options()]


assign_primary_coaches()
TRAINING_PLANS.extend(DEMO_TRAINING_PLANS)
TRAINING_PLANS.extend(ADDITIONAL_DEMO_TRAINING_PLANS)
FITNESS_TESTS.extend(DEMO_FITNESS_TESTS)
FITNESS_TESTS.extend(ADDITIONAL_DEMO_FITNESS_TESTS)
INJURY_RECORDS.extend(DEMO_INJURY_RECORDS)
INJURY_RECORDS.extend(ADDITIONAL_DEMO_INJURY_RECORDS)
MATCH_RESULTS.extend(DEMO_MATCH_RESULTS)
MATCH_RESULTS.extend(ADDITIONAL_DEMO_MATCH_RESULTS)
PLAN_ID_COUNTER = max(plan["id"] for plan in TRAINING_PLANS) + 1

ROLE_PERMISSION_MATRIX = [
    {"module": "运动员档案", "admin": "新增 / 修改 / 删除 / 查询", "coach": "查询与训练关联查看"},
    {"module": "训练数据录入", "admin": "导入 / 导出 / 编辑 / 删除", "coach": "新增训练与专项记录"},
    {"module": "体能测试", "admin": "全量维护与异常处理", "coach": "录入、查询、统计"},
    {"module": "伤病与康复", "admin": "严重伤病确认、归档、复训审批", "coach": "登记轻中度伤病与复诊记录"},
    {"module": "比赛成绩", "admin": "维护成绩库与阶段报告", "coach": "查询、复盘、提交分析"},
    {"module": "用户权限 / 系统配置", "admin": "账号、角色、字典、数据库账号", "coach": "不可访问"},
]

DATABASE_ACCOUNTS = [
    {"username": "admin_app", "role": "管理员", "privilege": "ALL PRIVILEGES", "scope": "pingpang_db.*"},
    {"username": "coach_app", "role": "教练员", "privilege": "SELECT, INSERT, UPDATE, DELETE", "scope": "业务数据表"},
]

SYSTEM_PARAMETERS = [
    {"name": "系统名称", "value": "乒乓球运动员综合训练监控管理系统", "owner": "管理员"},
    {"name": "数据库名", "value": "pingpang_db", "owner": "数据库设计组"},
    {"name": "默认训练周期", "value": "按周 / 月 / 赛前阶段统计", "owner": "训练模块"},
    {"name": "异常提示规则", "value": "伤病严重、体能低分、康复逾期触发预警", "owner": "测试与运维"},
]

DICTIONARY_GROUPS = [
    {"name": "运动等级", "items": ["二级运动员", "一级运动员", "国家级", "健将级", "青年队"]},
    {"name": "训练强度", "items": list(INTENSITY_LABELS.keys())},
    {"name": "步法类型", "items": []},
    {"name": "技战术类型", "items": []},
    {"name": "落点分布", "items": []},
    {"name": "伤病程度", "items": INJURY_SEVERITY_OPTIONS},
    {"name": "恢复状态", "items": INJURY_RECOVERY_STATUS_OPTIONS},
    {"name": "比赛结果", "items": ["胜", "负", "平"]},
]
refresh_dictionary_groups()

DATABASE_HEALTH_CHECKS = [
    {"item": "数据表", "target": "不少于 6 张核心表", "status": "8 张表已覆盖", "class": "success"},
    {"item": "约束", "target": "主键、外键、非空、唯一约束", "status": "SQL 脚本已配置", "class": "success"},
    {"item": "索引", "target": "姓名、日期组合查询优化", "status": "按课程要求保留检查项", "class": "warning"},
    {"item": "视图", "target": "运动员综合档案、月度训练汇总", "status": "需在最终 SQL 中验收", "class": "warning"},
    {"item": "存储过程", "target": "批量导入月度训练数据、按等级筛选", "status": "需在最终 SQL 中验收", "class": "warning"},
    {"item": "触发器", "target": "新增伤病后自动刷新运动员伤病状态", "status": "伤病模块已实现同等业务联动", "class": "success"},
    {"item": "事务", "target": "训练数据 + 体能数据同步提交", "status": "体能模块已提供回滚逻辑", "class": "success"},
]

MODULE_FEATURES = {
    "训练计划管理": [
        {"title": "训练周期", "desc": "按周、月、赛前周期安排训练目标与重点。"},
        {"title": "计划执行", "desc": "跟踪训练完成情况、负荷变化和教练反馈。"},
        {"title": "计划调整", "desc": "根据运动员状态及时调整训练内容和强度。"},
        {"title": "计划归档", "desc": "沉淀历史训练计划，便于复盘与对比。"},
    ],
    "体能测试评估": [
        {"title": "测试记录", "desc": "录入速度、力量、耐力、灵敏等体能指标。"},
        {"title": "等级评估", "desc": "结合队内标准形成体能等级和风险提示。"},
        {"title": "趋势分析", "desc": "跟踪单项体能指标的长期变化。"},
        {"title": "训练建议", "desc": "辅助教练制定针对性体能提升方案。"},
    ],
    "伤病记录管理": [
        {"title": "伤病档案", "desc": "记录伤病部位、时间、程度和处理方案。"},
        {"title": "风险标记", "desc": "对重点关注运动员进行健康风险提示。"},
        {"title": "复诊跟踪", "desc": "维护复查、治疗和恢复过程记录。"},
        {"title": "历史查询", "desc": "快速查看运动员过往伤病情况。"},
    ],
    "康复跟踪预警": [
        {"title": "康复计划", "desc": "安排阶段性康复目标和训练限制。"},
        {"title": "恢复进度", "desc": "记录疼痛、活动度、训练适应等恢复指标。"},
        {"title": "预警提醒", "desc": "对异常恢复进度和复发风险进行提示。"},
        {"title": "复训评估", "desc": "辅助判断运动员是否适合恢复正常训练。"},
    ],
    "比赛成绩报告": [
        {"title": "成绩记录", "desc": "维护比赛时间、对手、比分和名次。"},
        {"title": "技战术复盘", "desc": "总结发接发、相持、关键分等表现。"},
        {"title": "对手分析", "desc": "沉淀重点对手特征和历史交手记录。"},
        {"title": "报告输出", "desc": "形成面向教练组的阶段性比赛报告。"},
    ],
    "系统配置": [
        {"title": "数据字典", "desc": "统一维护运动等级、训练强度、伤病状态等选项。"},
        {"title": "基础参数", "desc": "配置系统名称、训练周期、导入字段等基础信息。"},
        {"title": "导入模板", "desc": "维护 Excel 批量导入模板和字段规则。"},
        {"title": "系统维护", "desc": "支持基础配置检查和运行状态维护。"},
    ],
    "新增运动员": [
        {"title": "基础信息", "desc": "录入姓名、性别、年龄、联系方式等基础档案。"},
        {"title": "竞技信息", "desc": "维护运动等级、打法、持拍手和专项特点。"},
        {"title": "健康状态", "desc": "同步初始健康状态与重点关注标签。"},
        {"title": "档案提交", "desc": "提交后进入运动员档案统一管理。"},
    ],
    "编辑运动员": [
        {"title": "档案更新", "desc": "更新运动员基础资料和竞技信息。"},
        {"title": "状态维护", "desc": "调整当前训练、健康和参赛状态。"},
        {"title": "记录关联", "desc": "关联训练、伤病、体能和比赛数据。"},
        {"title": "变更确认", "desc": "保存后同步更新档案列表。"},
    ],
    "历史伤病": [
        {"title": "伤病时间线", "desc": "查看运动员历次伤病发生和恢复情况。"},
        {"title": "治疗记录", "desc": "追踪处理方案、复诊结论和训练限制。"},
        {"title": "复发风险", "desc": "辅助识别长期高风险部位。"},
        {"title": "归档查询", "desc": "为训练计划调整提供依据。"},
    ],
}


def env_int(name, default):
    try:
        return int(os.getenv(name, default))
    except (TypeError, ValueError):
        return default


def create_app():
    from coaches import bp as coaches_bp

    app = Flask(__name__)
    app.config["SECRET_KEY"] = (
        os.getenv("SECRET_KEY") or os.getenv("FLASK_SECRET_KEY") or secrets.token_hex(32)
    )
    app.config["MAX_CONTENT_LENGTH"] = env_int("MAX_CONTENT_LENGTH", 5 * 1024 * 1024)

    @app.context_processor
    def inject_layout_data():
        user = current_user()
        role = user["role"] if user else None
        visible_nav_groups, visible_nav_items = build_navigation_for_role(role) if role else ([], [])
        return {
            "current_user": user,
            "nav_groups": visible_nav_groups,
            "nav_items": visible_nav_items,
            "csrf_token": csrf_token,
        }

    @app.before_request
    def require_login():
        public_endpoints = {"login", "static", "healthz"}
        if request.endpoint in public_endpoints:
            return None
        if not session.get("username"):
            next_url = request.full_path if request.query_string else request.path
            return redirect(url_for("login", next=next_url))
        return None

    @app.before_request
    def protect_write_requests():
        return validate_csrf_token()

    @app.route("/healthz")
    def healthz():
        return {"status": "ok"}

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if session.get("username"):
            return redirect(url_for("index"))

        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "").strip()
            user = USERS.get(username)

            if user and user["password"] == password:
                session["username"] = username
                flash(f"欢迎回来，{user['name']}。", "success")
                requested_next = request.args.get("next")
                next_url = requested_next if is_safe_redirect_url(requested_next) else url_for("matches.list")
                return redirect(next_url)

            flash("用户名或密码错误，请重新输入。", "danger")

        return render_template("auth/login.html")

    @app.route("/logout")
    def logout():
        session.clear()
        flash("您已安全退出系统。", "success")
        return redirect(url_for("login"))

    @app.route("/")
    @role_required("admin", "coach")
    def index():
        return redirect(url_for("matches.list"))

    players_bp = Blueprint("players", __name__, url_prefix="/players")

    @players_bp.route("/", endpoint="list")
    @role_required("admin", "coach")
    def players_list():
        players, active_condition_count = filter_players(request.args)
        page_size = 10
        page = request.args.get("page", "1")
        page = int(page) if page.isdigit() and int(page) > 0 else 1
        filtered_count = len(players)
        total_pages = max(1, (filtered_count + page_size - 1) // page_size)
        page = min(page, total_pages)
        page_start = (page - 1) * page_size
        paged_players = players[page_start : page_start + page_size]
        pagination_args = request.args.to_dict(flat=True)
        pagination_args.pop("page", None)
        return render_template(
            "players/list.html",
            players=paged_players,
            total_count=len(PLAYERS),
            filtered_count=filtered_count,
            active_condition_count=active_condition_count,
            logic=request.args.get("logic", "and"),
            level_labels=PLAYER_LEVEL_LABELS,
            injury_status_labels=PLAYER_INJURY_STATUS_LABELS,
            gender_options=PLAYER_GENDER_OPTIONS,
            pagination={
                "page": page,
                "page_size": page_size,
                "total_pages": total_pages,
                "has_prev": page > 1,
                "has_next": page < total_pages,
                "prev_page": page - 1,
                "next_page": page + 1,
                "start": page_start + 1 if filtered_count else 0,
                "end": min(page_start + page_size, filtered_count),
                "args": pagination_args,
            },
        )

    @players_bp.route("/create", methods=["GET", "POST"], endpoint="create")
    @role_required("admin")
    def players_create():
        if request.method == "POST":
            try:
                player = validate_player_form(request.form)
                if any(item["student_no"] == player["student_no"] for item in PLAYERS):
                    raise ValidationError("学号已存在，请更换后再提交。")
                PLAYERS.append({"id": next_id(PLAYERS), **player})
                flash("运动员档案创建成功。", "success")
                return redirect(url_for("players.list"))
            except ValidationError as exc:
                flash(str(exc), "danger")

        return render_template(
            "players/create.html",
            level_labels=PLAYER_LEVEL_LABELS,
            injury_status_labels=PLAYER_INJURY_STATUS_LABELS,
            gender_options=PLAYER_GENDER_OPTIONS,
            coaches=COACHES,
        )

    @players_bp.route("/<int:player_id>", endpoint="detail")
    @role_required("admin", "coach")
    def players_detail(player_id):
        player = find_player(player_id)
        if not player:
            flash("运动员不存在。", "warning")
            return redirect(url_for("players.list"))
        return render_template(
            "players/detail.html",
            player=player,
            training_plans=[item for item in TRAINING_PLANS if item["athlete_id"] == player_id],
            footwork_records=[
                enrich_footwork_record(item)
                for item in training_repo.list_footwork_records_by_athlete(player_id)
            ],
            technique_records=[
                enrich_technique_tactic_record(item)
                for item in training_repo.list_technique_records_by_athlete(player_id)
            ],
            fitness_tests=[
                enrich_fitness_record(item)
                for item in FITNESS_TESTS
                if item["athlete_id"] == player_id
            ],
            injury_records=[
                enrich_injury_record(item)
                for item in INJURY_RECORDS
                if item["athlete_id"] == player_id and not item.get("is_deleted")
            ],
            match_records=[
                enrich_match_record(item)
                for item in MATCH_RESULTS
                if item["athlete_id"] == player_id
            ],
        )

    @players_bp.route("/<int:player_id>/edit", methods=["GET", "POST"], endpoint="edit")
    @role_required("admin")
    def players_edit(player_id):
        player = find_player(player_id)
        if not player:
            flash("运动员不存在。", "warning")
            return redirect(url_for("players.list"))

        if request.method == "POST":
            try:
                validated = validate_player_form(request.form)
                if any(
                    item["id"] != player_id and item["student_no"] == validated["student_no"]
                    for item in PLAYERS
                ):
                    raise ValidationError("学号已被其他运动员使用。")
                player.update(validated)
                flash("运动员档案更新成功。", "success")
                return redirect(url_for("players.detail", player_id=player_id))
            except ValidationError as exc:
                flash(str(exc), "danger")

        return render_template(
            "players/edit.html",
            player=player,
            level_labels=PLAYER_LEVEL_LABELS,
            injury_status_labels=PLAYER_INJURY_STATUS_LABELS,
            gender_options=PLAYER_GENDER_OPTIONS,
            coaches=COACHES,
        )

    @players_bp.route("/<int:player_id>/delete", methods=["POST"], endpoint="delete")
    @role_required("admin")
    def players_delete(player_id):
        player = find_player(player_id)
        if not player:
            flash("运动员不存在。", "warning")
        else:
            PLAYERS.remove(player)
            flash("运动员档案已删除。", "success")
        return redirect(url_for("players.list"))

    training_bp = Blueprint("training", __name__, url_prefix="/training")

    @training_bp.route("/plans", methods=["GET", "POST"], endpoint="plans")
    @role_required("admin", "coach")
    def training_plans():
        """训练计划列表、查询、新增"""
        source_match = None
        source_analysis = None
        source_analysis_id = ""
        source_notice = ""
        source_type = ""
        source_match_id = ""
        source_summary = ""
        default_athlete_id = request.args.get("athlete_id", "").strip()
        default_coach_id = request.args.get("coach_id", "").strip()

        if request.method == "GET":
            source_type = request.args.get("source_type", "").strip()
            source_match_id = request.args.get("source_match_id", "").strip()
            if source_type == "match" and source_match_id.isdigit():
                source_match = get_match_record(int(source_match_id))
                if source_match:
                    default_athlete_id = str(source_match["athlete_id"])
                    source_analysis = get_latest_confirmed_analysis_for_match(source_match["id"])
                    if source_analysis:
                        source_type = "match_analysis"
                        source_analysis_id = str(source_analysis["id"])
                        source_summary = build_match_analysis_plan_source_summary(source_analysis)
                    else:
                        source_summary = build_match_plan_source_summary(source_match)
                        source_notice = "暂无已确认技战术分析，仅可作为比赛结果来源；需先完成分析确认后再作为普通改进计划依据。"

        if request.method == "POST":
            athlete_id = request.form.get("athlete_id")
            coach_id = request.form.get("coach_id")
            plan_datetime = request.form.get("plan_datetime")
            content = request.form.get("content")
            intensity = request.form.get("intensity")
            duration = request.form.get("duration")
            location = request.form.get("location")
            plan_status = request.form.get("plan_status", "draft").strip() or "draft"
            source_type = request.form.get("source_type", "").strip()
            source_match_id = request.form.get("source_match_id", "").strip()
            source_analysis_id = request.form.get("source_analysis_id", "").strip()
            source_summary = request.form.get("source_summary", "").strip()
            plan_items, item_errors = parse_training_plan_items(request.form)

            if source_type == "match_analysis" and source_analysis_id.isdigit():
                source_analysis = get_match_analysis(int(source_analysis_id))
                if source_analysis and source_analysis["status"] == "confirmed":
                    source_match = get_match_record(source_analysis["match_id"])
                    source_match_id = str(source_analysis["match_id"])
                    if not source_summary:
                        source_summary = build_match_analysis_plan_source_summary(source_analysis)
                else:
                    item_errors.append("只有已确认技战术分析可以作为普通改进训练计划来源。")
            elif source_type == "match" and source_match_id.isdigit():
                source_match = get_match_record(int(source_match_id))
                if source_match and not source_summary:
                    source_summary = build_match_plan_source_summary(source_match)

            # 🔽 调用校验函数
            errors = validate_training_plan_data(athlete_id, coach_id, plan_datetime, content, intensity, duration)
            errors.extend(item_errors)
            if plan_status not in TRAINING_PLAN_STATUS_LABELS:
                errors.append("训练计划状态非法。")
            else:
                errors.extend(validate_training_plan_status_choice(plan_status, plan_items))
            if (
                source_type == "match_analysis"
                and source_match
                and athlete_id
                and athlete_id.isdigit()
                and source_match["athlete_id"] != int(athlete_id)
            ):
                errors.append("训练计划运动员必须与来源比赛运动员一致。")
            
            if errors:
                for err in errors:
                    flash(err, "danger")
                return redirect(url_for("training.plans"))   # 直接重定向回列表页

            athlete = next((p for p in PLAYERS if str(p["id"]) == athlete_id), None)
            coach = next((c for c in COACHES if str(c["id"]) == coach_id), None)
            if not athlete or not coach:
                flash("运动员或教练不存在", "danger")
                return redirect(url_for("training.plans"))

            global PLAN_ID_COUNTER, PLAN_ITEM_ID_COUNTER
            new_plan_id = PLAN_ID_COUNTER
            new_plan = {
                        "id": new_plan_id,
                        "athlete_id": int(athlete_id),
                        "athlete_name": athlete["name"],
                        "coach_id": int(coach_id),
                        "coach_name": coach["name"],
                        "plan_datetime": plan_datetime,
                        "content": content,
                        "intensity": intensity,
                        "duration": int(duration) if duration else 60,
                        "location": location or "",
                        "status": plan_status,
                        "status_label": TRAINING_PLAN_STATUS_LABELS[plan_status],
                        "source_type": source_type,
                        "source_match_id": int(source_match_id) if source_match_id.isdigit() else None,
                        "source_analysis_id": int(source_analysis_id) if source_analysis_id.isdigit() else None,
                        "source_summary": source_summary,
                    }
            TRAINING_PLANS.append(new_plan)
            for item in plan_items:
                TRAINING_PLAN_ITEMS.append(
                    {
                        "id": PLAN_ITEM_ID_COUNTER,
                        "plan_id": new_plan_id,
                        **item,
                    }
                )
                PLAN_ITEM_ID_COUNTER += 1
            PLAN_ID_COUNTER += 1
            flash("训练计划添加成功", "success")
            return redirect(url_for("training.plans"))

        # GET 请求：查询与列表展示
        athlete_name = request.args.get("athlete_name", "").strip()
        start_date = request.args.get("start_date", "").strip()
        end_date = request.args.get("end_date", "").strip()
        content_keyword = request.args.get("content", "").strip()

        filtered = TRAINING_PLANS.copy()
        if athlete_name:
            filtered = [p for p in filtered if athlete_name.lower() in p["athlete_name"].lower()]
        if start_date:
            filtered = [p for p in filtered if p["plan_datetime"] >= start_date]
        if end_date:
            filtered = [p for p in filtered if p["plan_datetime"] <= end_date]
        if content_keyword:
            filtered = [p for p in filtered if content_keyword.lower() in p["content"].lower()]

        filtered.sort(key=lambda x: x["plan_datetime"], reverse=True)

        return render_template(
            "training/plans.html",
            plans=filtered,
            athletes=PLAYERS,
            coaches=COACHES,
            athlete_name=athlete_name,
            start_date=start_date,
            end_date=end_date,
            content=content_keyword,
            default_athlete_id=default_athlete_id,
            default_coach_id=default_coach_id,
            source_match=source_match,
            source_analysis=source_analysis,
            source_analysis_id=source_analysis_id,
            source_notice=source_notice,
            source_type=source_type,
            source_match_id=source_match_id,
            source_summary=source_summary,
            plan_items_by_plan_id=group_training_plan_items_by_plan(),
            item_module_options=[
                {"value": value, "label": label}
                for value, label in TRAINING_PLAN_ITEM_MODULE_LABELS.items()
            ],
            plan_status_options=[
                {"value": value, "label": label}
                for value, label in TRAINING_PLAN_STATUS_LABELS.items()
            ],
            plan_status_labels=TRAINING_PLAN_STATUS_LABELS,
        )
   
    @training_bp.route("/plans/<int:plan_id>/edit", methods=["GET", "POST"])
    @role_required("admin", "coach")
    def edit_plan(plan_id):
        """编辑训练计划"""
        plan = next((p for p in TRAINING_PLANS if p["id"] == plan_id), None)
        if not plan:
            flash("训练计划不存在", "danger")
            return redirect(url_for("training.plans"))

        if request.method == "POST":
            athlete_id = request.form.get("athlete_id")
            coach_id = request.form.get("coach_id")
            plan_datetime = request.form.get("plan_datetime")
            content = request.form.get("content")
            intensity = request.form.get("intensity")
            duration = request.form.get("duration")
            location = request.form.get("location")
            plan_status = request.form.get("plan_status", plan.get("status", "published")).strip() or plan.get("status", "published")

            # 🔽 调用统一的校验函数
            errors = validate_training_plan_data(athlete_id, coach_id, plan_datetime, content, intensity, duration)
            if plan_status not in TRAINING_PLAN_STATUS_LABELS:
                errors.append("训练计划状态非法。")
            else:
                errors.extend(validate_training_plan_status_choice(plan_status, get_training_plan_items(plan["id"])))
            if errors:
                for err in errors:
                    flash(err, "danger")
                return render_template(
                    "training/plan_form.html",
                    plan=plan,
                    athletes=PLAYERS,
                    coaches=COACHES,
                    plan_status_options=[
                        {"value": value, "label": label}
                        for value, label in TRAINING_PLAN_STATUS_LABELS.items()
                    ],
                    plan_status_labels=TRAINING_PLAN_STATUS_LABELS,
                )

            # 查找运动员和教练（保留原有逻辑）
            athlete = next((p for p in PLAYERS if str(p["id"]) == athlete_id), None)
            coach = next((c for c in COACHES if str(c["id"]) == coach_id), None)
            if not athlete or not coach:
                flash("运动员或教练不存在", "danger")
                return render_template(
                    "training/plan_form.html",
                    plan=plan,
                    athletes=PLAYERS,
                    coaches=COACHES,
                    plan_status_options=[
                        {"value": value, "label": label}
                        for value, label in TRAINING_PLAN_STATUS_LABELS.items()
                    ],
                    plan_status_labels=TRAINING_PLAN_STATUS_LABELS,
                )

            # 更新计划（保留原有逻辑）
            plan.update({
                "athlete_id": int(athlete_id),
                "athlete_name": athlete["name"],
                "coach_id": int(coach_id),
                "coach_name": coach["name"],
                "plan_datetime": plan_datetime,
                "content": content,
                "intensity": intensity,
                "duration": int(duration) if duration else 60,
                "location": location or "",
                "status": plan_status,
            })
            flash("训练计划更新成功", "success")
            return redirect(url_for("training.plans"))

        return render_template(
            "training/plan_form.html",
            plan=plan,
            athletes=PLAYERS,
            coaches=COACHES,
            plan_status_options=[
                {"value": value, "label": label}
                for value, label in TRAINING_PLAN_STATUS_LABELS.items()
            ],
            plan_status_labels=TRAINING_PLAN_STATUS_LABELS,
        )


    @training_bp.route("/plans/<int:plan_id>/delete", methods=["POST"])
    @role_required("admin", "coach")
    def delete_plan(plan_id):
        """删除训练计划"""
        plan = next((p for p in TRAINING_PLANS if p["id"] == plan_id), None)
        if not plan:
            flash("训练计划不存在", "danger")
            return redirect(url_for("training.plans"))
        if not can_delete_training_plan(plan, current_user()):
            return render_template("auth/forbidden.html"), 403
        TRAINING_PLANS[:] = [p for p in TRAINING_PLANS if p["id"] != plan_id]
        record_audit_log(
            action="delete",
            target_type="training_plan",
            target_id=plan_id,
            before=plan,
            user=current_user(),
        )
        flash("训练计划已删除", "success")
        return redirect(url_for("training.plans"))
    @training_bp.route("/footwork", methods=["GET", "POST"], endpoint="footwork")
    @role_required("admin", "coach")
    def training_footwork():
        editing_record = get_editing_footwork_record(request.args.get("edit_id", "").strip())
        if request.method == "POST":
            try:
                if editing_record:
                    save_footwork_training_record(request.form, current_user()["username"], editing_record["id"])
                    flash("步法训练记录已更新。", "success")
                else:
                    save_footwork_training_record(request.form, current_user()["username"])
                    flash("步法训练记录已保存。", "success")
                return redirect(url_for("training.footwork"))
            except ValidationError as exc:
                flash(str(exc), "danger")

        records, active_condition_count = filter_footwork_training_records(request.args)
        page_size = 10
        page = request.args.get("page", "1")
        page = int(page) if page.isdigit() and int(page) > 0 else 1
        filtered_count = len(records)
        total_pages = max(1, (filtered_count + page_size - 1) // page_size)
        page = min(page, total_pages)
        page_start = (page - 1) * page_size
        paged_records = records[page_start : page_start + page_size]
        pagination_args = request.args.to_dict(flat=True)
        pagination_args.pop("page", None)
        return render_template(
            "training/footwork.html",
            records=paged_records,
            athletes=PLAYERS,
            editing_record=editing_record,
            footwork_options=get_footwork_options(),
            total_count=training_repo.count_footwork_records(),
            filtered_count=filtered_count,
            active_condition_count=active_condition_count,
            module_plan_items=build_execution_frame_items("footwork"),
            pagination={
                "page": page,
                "page_size": page_size,
                "total_pages": total_pages,
                "has_prev": page > 1,
                "has_next": page < total_pages,
                "prev_page": page - 1,
                "next_page": page + 1,
                "start": page_start + 1 if filtered_count else 0,
                "end": min(page_start + page_size, filtered_count),
                "args": pagination_args,
            },
        )

    @training_bp.route("/batch-import", methods=["GET", "POST"], endpoint="batch_import")
    @role_required("admin", "coach")
    def legacy_batch_import_redirect():
        if request.method == "POST":
            try:
                editing_record = get_editing_footwork_record(request.args.get("edit_id", "").strip())
                record_id = editing_record["id"] if editing_record else None
                save_footwork_training_record(request.form, current_user()["username"], record_id)
                flash("步法训练记录已保存。", "success")
            except ValidationError as exc:
                flash(str(exc), "danger")
            return redirect(url_for("training.footwork"))
        return redirect(url_for("training.footwork", edit_id=request.args.get("edit_id", "")))

    @training_bp.route("/technique-tactic", methods=["GET", "POST"], endpoint="technique_tactic")
    @role_required("admin", "coach")
    def training_technique_tactic():
        editing_record = get_editing_technique_tactic_record(request.args.get("edit_id", "").strip())
        if request.method == "POST":
            try:
                if editing_record:
                    save_technique_tactic_record(request.form, current_user()["username"], editing_record["id"])
                    flash("技战术训练记录已更新。", "success")
                else:
                    save_technique_tactic_record(request.form, current_user()["username"])
                    flash("技战术训练记录已保存。", "success")
                return redirect(url_for("training.technique_tactic"))
            except ValidationError as exc:
                flash(str(exc), "danger")

        records, active_condition_count = filter_technique_tactic_records(request.args)
        page_size = 10
        page = request.args.get("page", "1")
        page = int(page) if page.isdigit() and int(page) > 0 else 1
        filtered_count = len(records)
        total_pages = max(1, (filtered_count + page_size - 1) // page_size)
        page = min(page, total_pages)
        page_start = (page - 1) * page_size
        paged_records = records[page_start : page_start + page_size]
        pagination_args = request.args.to_dict(flat=True)
        pagination_args.pop("page", None)
        return render_template(
            "training/technique_tactic.html",
            records=paged_records,
            athletes=PLAYERS,
            editing_record=editing_record,
            total_count=training_repo.count_technique_tactic_records(),
            filtered_count=filtered_count,
            active_condition_count=active_condition_count,
            technique_categories=get_technique_categories(),
            technique_options=get_technique_options(),
            landing_options=get_landing_options(),
            landing_concentration_options=LANDING_CONCENTRATION_OPTIONS,
            serve_frequency_options=SERVE_FREQUENCY_OPTIONS,
            module_plan_items=build_execution_frame_items("technique_tactic"),
            pagination={
                "page": page,
                "page_size": page_size,
                "total_pages": total_pages,
                "has_prev": page > 1,
                "has_next": page < total_pages,
                "prev_page": page - 1,
                "next_page": page + 1,
                "start": page_start + 1 if filtered_count else 0,
                "end": min(page_start + page_size, filtered_count),
                "args": pagination_args,
            },
        )

    @training_bp.route("/records", endpoint="record")
    @role_required("admin", "coach")
    def legacy_training_records_redirect():
        return redirect(url_for("training.technique_tactic", **request.args), code=301)

    @training_bp.route("/training_record", endpoint="training_record")
    @role_required("admin", "coach")
    def legacy_training_record_redirect():
        return redirect(url_for("training.technique_tactic", **request.args), code=301)

    @training_bp.route("/footwork/<int:record_id>/edit", methods=["POST"], endpoint="edit_footwork")
    @role_required("admin", "coach")
    def edit_footwork_record(record_id):
        try:
            save_footwork_training_record(request.form, current_user()["username"], record_id)
            flash("步法训练记录已更新。", "success")
        except ValidationError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("training.footwork", edit_id=record_id))
        return redirect(url_for("training.footwork"))

    @training_bp.route("/technique-tactic/<int:record_id>/edit", methods=["POST"], endpoint="edit_technique_tactic")
    @role_required("admin", "coach")
    def edit_technique_tactic_record(record_id):
        try:
            save_technique_tactic_record(request.form, current_user()["username"], record_id)
            flash("技战术训练记录已更新。", "success")
        except ValidationError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("training.technique_tactic", edit_id=record_id))
        return redirect(url_for("training.technique_tactic"))

    @training_bp.route("/records/<int:record_id>/edit", methods=["POST"], endpoint="edit_record")
    @role_required("admin", "coach")
    def legacy_edit_training_record(record_id):
        return edit_technique_tactic_record(record_id)

    @training_bp.route("/footwork/<int:record_id>/delete", methods=["POST"], endpoint="delete_footwork")
    @role_required("admin", "coach")
    def delete_footwork_record(record_id):
        if not training_repo.delete_footwork_record(record_id):
            flash("步法训练记录不存在。", "danger")
        else:
            flash("步法训练记录已删除。", "success")
        return redirect(url_for("training.footwork"))

    @training_bp.route("/technique-tactic/<int:record_id>/delete", methods=["POST"], endpoint="delete_technique_tactic")
    @role_required("admin", "coach")
    def delete_technique_tactic_record(record_id):
        if not training_repo.delete_technique_tactic_record(record_id):
            flash("技战术训练记录不存在。", "danger")
        else:
            flash("技战术训练记录已删除。", "success")
        return redirect(url_for("training.technique_tactic"))

    @training_bp.route("/records/<int:record_id>/delete", methods=["POST"], endpoint="delete_record")
    @role_required("admin", "coach")
    def legacy_delete_training_record(record_id):
        return delete_technique_tactic_record(record_id)

    @training_bp.route("/footwork/import-excel", methods=["POST"], endpoint="import_footwork_excel")
    @role_required("admin", "coach")
    def import_footwork_excel():
        file = request.files.get("training_excel")
        if not file:
            flash("请先选择 Excel 文件。", "warning")
            return redirect(url_for("training.footwork"))

        imported_count, error_rows = import_footwork_training_excel(file, current_user()["username"])
        if error_rows:
            flash(f"成功导入 {imported_count} 条，跳过 {len(error_rows)} 条异常数据。", "warning")
            for err in error_rows[:5]:
                flash(err, "warning")
        else:
            flash(f"成功导入 {imported_count} 条步法训练记录。", "success")

        return redirect(url_for("training.footwork"))

    @training_bp.route("/records/import-excel", methods=["POST"], endpoint="import_skill_excel")
    @role_required("admin", "coach")
    def legacy_import_skill_excel():
        return import_footwork_excel()

    @training_bp.route("/technique-tactic/export", methods=["GET"], endpoint="export_technique_tactic")
    @role_required("admin", "coach")
    def export_technique_tactic_records():
        records, _ = filter_technique_tactic_records(request.args)
        wb = openpyxl.Workbook()
        ws = wb.active
        write_technique_tactic_sheet(ws, records)
        return send_workbook(wb, "technique_tactic_training_records.xlsx")

    @training_bp.route("/records/export", methods=["GET"], endpoint="export_records")
    @role_required("admin", "coach")
    def legacy_export_training_records():
        return export_technique_tactic_records()

    @training_bp.route("/import-excel", methods=["POST"], endpoint="import_excel")
    @role_required("admin", "coach")
    def training_import_excel():
        file = request.files.get("training_excel")
        if not file:
            flash("请先选择 Excel 文件。", "warning")
            return redirect(url_for("training.plans"))

        from datetime import datetime, timedelta
        error_rows = []
        imported_count = 0
        global PLAN_ID_COUNTER

        def plan_identity(plan):
            normalize_text = lambda value: " ".join(str(value or "").split())
            return (
                plan["athlete_id"],
                plan["coach_id"],
                plan["plan_datetime"],
                normalize_text(plan["content"]),
                plan["intensity"],
                plan["duration"],
                normalize_text(plan.get("location")),
            )

        existing_plan_identities = {plan_identity(plan) for plan in TRAINING_PLANS}
        expected_header_aliases = {
            "athlete_name": {"运动员", "运动员姓名"},
            "coach_name": {"教练", "教练员", "教练姓名"},
            "plan_datetime": {"训练日期", "训练时间", "计划时间"},
            "content": {"内容", "训练内容", "计划内容"},
            "intensity": {"强度", "训练强度"},
            "duration": {"时长", "时长(分钟)", "训练时长", "训练时长(分钟)"},
            "location": {"地点", "训练地点"},
        }
        footwork_record_headers = {"运动员学号", "训练日期", "步法类型", "训练时长(分钟)", "训练组数"}
        legacy_skill_headers = {"运动员", "训练日期", "步法训练", "击球技术", "多球时长", "训练强度", "备注"}

        def normalize_header(value):
            return str(value or "").strip().replace(" ", "")

        def resolve_training_plan_columns(ws):
            headers = [normalize_header(cell.value) for cell in ws[1]]
            header_set = {header for header in headers if header}
            if footwork_record_headers.issubset(header_set) or legacy_skill_headers.issubset(header_set):
                raise ValueError("该文件是步法训练记录模板，请在“步法训练”页面导入。")

            column_indexes = {}
            for field, aliases in expected_header_aliases.items():
                normalized_aliases = {normalize_header(alias) for alias in aliases}
                for column_index, header in enumerate(headers):
                    if header in normalized_aliases:
                        column_indexes[field] = column_index
                        break

            required_fields = ["athlete_name", "coach_name", "plan_datetime", "content"]
            if not all(field in column_indexes for field in required_fields):
                expected_headers = "运动员、教练、训练日期、内容、强度、时长(分钟)、地点"
                actual_headers = "、".join(header for header in headers if header) or "空表头"
                raise ValueError(f"训练计划 Excel 表头不匹配。期望包含：{expected_headers}；当前表头：{actual_headers}")
            return column_indexes

        def row_value(row, column_indexes, field):
            column_index = column_indexes.get(field)
            if column_index is None or column_index >= len(row):
                return None
            return row[column_index]

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            column_indexes = resolve_training_plan_columns(ws)

            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                athlete_name = row_value(row, column_indexes, "athlete_name")
                coach_name = row_value(row, column_indexes, "coach_name")
                plan_datetime = row_value(row, column_indexes, "plan_datetime")
                content = row_value(row, column_indexes, "content")
                intensity = row_value(row, column_indexes, "intensity")
                duration = row_value(row, column_indexes, "duration")
                location = row_value(row, column_indexes, "location")
                
                if not athlete_name and not content:
                    continue

                row_errors = []
                
                athlete = next((p for p in PLAYERS if p["name"] == athlete_name), None)
                coach = next((c for c in COACHES if c["name"] == coach_name), None)
                if not athlete:
                    row_errors.append(f"第{idx}行：找不到运动员 '{athlete_name}'")
                if not coach:
                    row_errors.append(f"第{idx}行：找不到教练 '{coach_name}'")
                
                plan_datetime_str = None
                try:
                    if isinstance(plan_datetime, datetime):
                        plan_datetime_str = plan_datetime.strftime("%Y-%m-%d %H:%M")
                    elif isinstance(plan_datetime, (int, float)):
                        plan_datetime_str = (
                            datetime(1899, 12, 30) + timedelta(days=plan_datetime)
                        ).strftime("%Y-%m-%d %H:%M")
                    else:
                        date_text = str(plan_datetime).strip()
                        for date_format in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
                            try:
                                plan_datetime_str = datetime.strptime(
                                    date_text, date_format
                                ).strftime("%Y-%m-%d %H:%M")
                                break
                            except ValueError:
                                continue
                        if plan_datetime_str is None:
                            raise ValueError("日期格式错误")
                except (TypeError, ValueError):
                    row_errors.append(f"第{idx}行：日期格式错误")

                valid_intensities = ["低", "中", "高", "极高"]
                if intensity and intensity not in valid_intensities:
                    row_errors.append(f"第{idx}行：强度 '{intensity}' 非法")

                try:
                    dur = int(duration) if duration else 60
                    if dur <= 0 or dur > 600:
                        row_errors.append(f"第{idx}行：时长必须为 1~600 分钟")
                except:
                    row_errors.append(f"第{idx}行：时长必须为数字")

                if row_errors:
                    error_rows.extend(row_errors)
                    continue

                new_plan = {
                    "id": PLAN_ID_COUNTER,
                    "athlete_id": athlete["id"],
                    "athlete_name": athlete["name"],
                    "coach_id": coach["id"],
                    "coach_name": coach["name"],
                    "plan_datetime": plan_datetime_str,
                    "content": content,
                    "intensity": intensity if intensity in valid_intensities else "中",
                    "duration": int(duration) if duration else 60,
                    "location": location or "",
                }
                new_plan_identity = plan_identity(new_plan)
                if new_plan_identity in existing_plan_identities:
                    error_rows.append(f"第{idx}行：重复训练计划已跳过")
                    continue
                TRAINING_PLANS.append(new_plan)
                existing_plan_identities.add(new_plan_identity)
                PLAN_ID_COUNTER += 1
                imported_count += 1

            if error_rows:
                flash(f"⚠️ 成功导入 {imported_count} 条，{len(error_rows)} 个错误已跳过：", "warning")
                for err in error_rows[:5]:
                    flash(f"  • {err}", "warning")
            else:
                flash(f"✅ 成功导入 {imported_count} 条训练计划！", "success")

        except ValueError as e:
            flash(f"❌ {str(e)}", "danger")
        except Exception as e:
            flash(f"❌ Excel 解析失败：{str(e)}", "danger")

        return redirect(url_for("training.plans"))
    @training_bp.route("/plans/export", methods=["GET"])
    @role_required("admin", "coach")
    def export_plans():
        """导出当前训练计划列表为 Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "训练计划"

        headers = ["运动员", "教练", "训练日期", "内容", "强度", "时长(分钟)", "地点"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row_idx, plan in enumerate(TRAINING_PLANS, 2):
            ws.cell(row=row_idx, column=1, value=plan["athlete_name"])
            ws.cell(row=row_idx, column=2, value=plan["coach_name"])
            ws.cell(row=row_idx, column=3, value=plan["plan_datetime"])
            ws.cell(row=row_idx, column=4, value=plan["content"])
            ws.cell(row=row_idx, column=5, value=plan["intensity"])
            ws.cell(row=row_idx, column=6, value=plan["duration"])
            ws.cell(row=row_idx, column=7, value=plan.get("location", ""))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="训练计划列表.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    injuries_bp = Blueprint("injuries", __name__, url_prefix="/injuries")

    @injuries_bp.route("/", methods=["GET", "POST"], endpoint="list")
    @role_required("admin", "coach")
    def injuries_list():
        if request.method == "POST":
            action = request.form.get("action", "save").strip()
            try:
                if action == "save":
                    save_injury_record(request.form, current_user())
                    flash("伤病记录已保存，运动员健康状态已联动刷新。", "success")
                elif action == "archive":
                    archive_injury_record(request.form, current_user())
                    flash("伤病记录已作废归档，运动员健康状态已重新计算。", "success")
                elif action == "followup":
                    save_injury_followup(request.form, current_user())
                    flash("复诊跟踪记录已保存。", "success")
                else:
                    raise ValidationError("未知操作类型，请刷新页面后重试。")
            except ValidationError as exc:
                flash(str(exc), "warning")
            except RuntimeError as exc:
                flash(f"伤病模块操作失败，已回滚本次变更：{exc}", "danger")
            return redirect(url_for("injuries.list", **build_injury_redirect_query(request.form)))

        injury_records, active_condition_count, query_errors = filter_injury_records(request.args)
        for error in query_errors:
            flash(error, "warning")
        editing_record = get_editing_injury_record(request.args.get("edit_id", "").strip())
        summary = build_injury_summary(injury_records)
        return render_template(
            "injuries/list.html",
            injury_records=injury_records,
            active_condition_count=active_condition_count,
            total_count=len(INJURY_RECORDS),
            editing_record=editing_record,
            athlete_choices=PLAYERS,
            severity_options=INJURY_SEVERITY_OPTIONS,
            recovery_status_options=INJURY_RECOVERY_STATUS_OPTIONS,
            summary=summary,
            training_alerts=build_training_alerts(injury_records),
        )

    @injuries_bp.route("/player/<int:player_id>/history", endpoint="history")
    @role_required("admin", "coach")
    def injuries_history(player_id):
        player = next((item for item in PLAYERS if item["id"] == player_id), None)
        if not player:
            flash("所选运动员不存在，无法查看伤病历史。", "warning")
            return redirect(url_for("players.list"))
        records = [
            enrich_injury_record(item)
            for item in INJURY_RECORDS
            if item["athlete_id"] == player_id
        ]
        records.sort(key=lambda item: (item["injury_date"], item["id"]), reverse=True)
        return render_template(
            "injuries/history.html",
            player=player,
            injury_records=records,
            summary=build_injury_summary(records),
            followups_by_record=group_followups_by_record(),
        )

    fitness_bp = Blueprint("fitness", __name__, url_prefix="/fitness")

    @fitness_bp.route("/training", endpoint="training")
    @role_required("admin", "coach")
    def fitness_training():
        items = build_execution_frame_items("fitness")
        return render_template(
            "fitness/training.html",
            pending_items=[item for item in items if item["status"] == "pending"],
            running_items=[item for item in items if item["status"] == "running"],
            completed_items=[item for item in items if item["status"] == "completed"],
        )

    @fitness_bp.route("/tests", methods=["GET", "POST"], endpoint="tests")
    @role_required("admin", "coach")
    def fitness_tests():
        if request.method == "POST":
            try:
                save_fitness_test(request.form, current_user()["username"])
                flash("体能测试记录已提交，训练数据已与体能数据同步写入。", "success")
            except ValidationError as exc:
                flash(str(exc), "warning")
            except RuntimeError as exc:
                flash(f"事务已回滚：{exc}", "danger")
            return redirect(url_for("fitness.tests", **build_redirect_query(request.form)))

        fitness_records, active_condition_count = filter_fitness_tests(request.args)
        editing_record = get_editing_fitness_record(request.args.get("edit_id", "").strip())
        summary = build_fitness_summary(fitness_records)
        return render_template(
            "fitness/tests.html",
            fitness_records=fitness_records,
            active_condition_count=active_condition_count,
            total_count=len(FITNESS_TESTS),
            editing_record=editing_record,
            athlete_choices=PLAYERS,
            coach_choices=COACHES,
            summary=summary,
            risk_options=[
                {"code": "stable", "label": "稳定"},
                {"code": "observe", "label": "观察"},
                {"code": "alert", "label": "预警"},
            ],
            intensity_options=INTENSITY_LABELS,
            plan_status_options=["进行中", "已完成", "已取消"],
        )

    rehab_bp = Blueprint("rehab", __name__, url_prefix="/rehab")

    @rehab_bp.route("/", endpoint="list")
    @role_required("admin", "coach")
    def rehab_list():
        rehab_records, active_condition_count = filter_rehab_records(request.args)
        summary = build_rehab_summary(rehab_records)
        return render_template(
            "rehab/list.html",
            rehab_records=rehab_records,
            active_condition_count=active_condition_count,
            total_count=len([item for item in INJURY_RECORDS if not item.get("is_deleted")]),
            summary=summary,
            risk_options=[
                {"code": "high", "label": "高风险"},
                {"code": "medium", "label": "观察"},
                {"code": "low", "label": "可控"},
            ],
            recovery_status_options=INJURY_RECOVERY_STATUS_OPTIONS,
        )

    matches_bp = Blueprint("matches", __name__, url_prefix="/matches")

    @matches_bp.route("/", methods=["GET", "POST"], endpoint="list")
    @role_required("admin", "coach")
    def matches_list():
        if request.method == "POST":
            try:
                save_match_record(request.form)
                if request.form.get("record_id", "").strip():
                    flash("比赛成绩已更新。", "success")
                else:
                    flash("比赛成绩已保存。", "success")
                return redirect(url_for("matches.list"))
            except ValidationError as exc:
                flash(str(exc), "danger")

        match_records, active_condition_count = filter_match_records(request.args)
        editing_record = get_editing_match_record(request.args.get("edit_id", "").strip())
        summary = build_match_summary(match_records)
        return render_template(
            "matches/list.html",
            match_records=match_records,
            active_condition_count=active_condition_count,
            total_count=len(MATCH_RESULTS),
            summary=summary,
            result_options=["胜", "负", "平"],
            match_type_options=["友谊赛", "正式比赛"],
            opponent_source_options=[
                {"value": "internal", "label": "系统内运动员"},
                {"value": "external", "label": "系统外参赛者"},
            ],
            editing_record=editing_record,
            athlete_choices=PLAYERS,
            analysis_records=build_match_analysis_view_models(),
            phase_options=[
                {"code": code, "label": label}
                for code, label in THREE_PHASE_LABELS.items()
            ],
        )

    @matches_bp.route("/<int:match_id>/analysis", methods=["POST"], endpoint="analysis")
    @role_required("admin", "coach")
    def matches_analysis(match_id):
        try:
            save_match_tactical_analysis(match_id, request.form, current_user()["username"])
            flash("比赛技战术分析已保存，当前标准待核验，未生成正式评级。", "success")
        except ValidationError as exc:
            flash(str(exc), "danger")
        return redirect(url_for("matches.list"))

    auth_bp = Blueprint("auth", __name__, url_prefix="/auth")

    @auth_bp.route("/users", endpoint="users")
    @role_required("admin")
    def auth_users():
        users = [
            {"username": username, **user}
            for username, user in USERS.items()
        ]
        summary = build_user_permission_summary(users)
        return render_template(
            "auth/users.html",
            users=users,
            summary=summary,
            permission_matrix=ROLE_PERMISSION_MATRIX,
            database_accounts=DATABASE_ACCOUNTS,
        )

    settings_bp = Blueprint("settings", __name__, url_prefix="/settings")

    @settings_bp.route("/dictionary", endpoint="dictionary")
    @role_required("admin")
    def settings_dictionary():
        return render_template(
            "settings/dictionary.html",
            parameters=SYSTEM_PARAMETERS,
            dictionary_groups=DICTIONARY_GROUPS,
            health_checks=DATABASE_HEALTH_CHECKS,
            import_templates=build_import_templates(),
        )

    stats_bp = Blueprint("stats", __name__, url_prefix="/stats")

    @stats_bp.route("/dashboard", endpoint="dashboard")
    @role_required("admin", "coach")
    def stats_dashboard():
        return render_template("stats/dashboard.html", stats=build_overall_stats())

    @stats_bp.route("/import-export", methods=["GET", "POST"], endpoint="import_export")
    @role_required("admin", "coach")
    def stats_import_export():
        if request.method == "POST":
            action = request.form.get("action", "")
            if action == "import_skill":
                file = request.files.get("skill_excel")
                if not file:
                    flash("请先选择步法训练 Excel 文件。", "warning")
                    return redirect(url_for("stats.import_export"))

                imported_count, error_rows = import_footwork_training_excel(file, current_user()["username"])
                if error_rows:
                    flash(f"成功导入 {imported_count} 条步法训练记录，跳过 {len(error_rows)} 条异常数据。", "warning")
                    for err in error_rows[:5]:
                        flash(err, "warning")
                else:
                    flash(f"成功导入 {imported_count} 条步法训练记录。", "success")
                return redirect(url_for("stats.import_export"))

        return render_template("stats/import_export.html")

    @stats_bp.route("/export/all", endpoint="export_all")
    @role_required("admin", "coach")
    def export_all_data():
        wb = openpyxl.Workbook()
        write_training_plan_sheet(wb.active, TRAINING_PLANS)
        write_footwork_sheet(wb.create_sheet("步法训练记录"), filter_footwork_training_records(request.args)[0])
        write_technique_tactic_sheet(wb.create_sheet("技战术训练记录"), filter_technique_tactic_records(request.args)[0])
        write_fitness_sheet(wb.create_sheet("体能测试记录"), FITNESS_TESTS)
        write_injury_sheet(wb.create_sheet("伤病记录"), INJURY_RECORDS)
        return send_workbook(wb, "table_tennis_training_all_data.xlsx")

    @stats_bp.route("/export/skill", endpoint="export_skill")
    @role_required("admin", "coach")
    def export_skill_data():
        records, _ = filter_footwork_training_records(request.args)
        wb = openpyxl.Workbook()
        write_footwork_sheet(wb.active, records)
        return send_workbook(wb, "footwork_training_records.xlsx")

    @stats_bp.route("/export/fitness", endpoint="export_fitness")
    @role_required("admin", "coach")
    def export_fitness_data():
        wb = openpyxl.Workbook()
        write_fitness_sheet(wb.active, FITNESS_TESTS)
        return send_workbook(wb, "fitness_tests.xlsx")

    app.register_blueprint(players_bp)
    app.register_blueprint(coaches_bp)
    app.register_blueprint(training_bp)
    app.register_blueprint(injuries_bp)
    app.register_blueprint(fitness_bp)
    app.register_blueprint(rehab_bp)
    app.register_blueprint(matches_bp)
    app.register_blueprint(auth_bp)
    app.register_blueprint(settings_bp)
    app.register_blueprint(stats_bp)

    return app


def filter_players(args):
    logic = args.get("logic", "and")
    predicates = []

    student_no = args.get("student_no", "").strip()
    name = args.get("name", "").strip()
    gender = args.get("gender", "").strip()
    level = args.get("level", "").strip()
    play_style = args.get("play_style", "").strip()
    injury_status = args.get("injury_status", "").strip()
    age_min = args.get("age_min", "").strip()
    age_max = args.get("age_max", "").strip()

    if student_no:
        predicates.append(lambda player, value=student_no: value in player["student_no"])
    if name:
        predicates.append(lambda player, value=name: value.lower() in player["name"].lower())
    if gender:
        predicates.append(lambda player, value=gender: player["gender"] == value)
    if level:
        predicates.append(lambda player, value=level: player["level_code"] == value)
    if play_style:
        predicates.append(lambda player, value=play_style: value.lower() in player["play_style"].lower())
    if injury_status:
        predicates.append(lambda player, value=injury_status: player["injury_status_code"] == value)
    if age_min.isdigit():
        predicates.append(lambda player, value=int(age_min): player["age"] >= value)
    if age_max.isdigit():
        predicates.append(lambda player, value=int(age_max): player["age"] <= value)

    if not predicates:
        return PLAYERS, 0

    if logic == "or":
        return [player for player in PLAYERS if any(check(player) for check in predicates)], len(predicates)

    return [player for player in PLAYERS if all(check(player) for check in predicates)], len(predicates)


def validate_player_form(form):
    student_no = form.get("student_no", "").strip()
    name = form.get("name", "").strip()
    gender = form.get("gender", "").strip()
    age = parse_int_range(form.get("age", "").strip(), "年龄", 1, 80)
    level_code = form.get("level_code", "").strip()
    play_style = form.get("play_style", "").strip()
    grip = form.get("grip", "").strip()
    contact_phone = form.get("contact_phone", "").strip()
    injury_status_code = form.get("injury_status_code", "healthy").strip()
    coach_id_text = form.get("coach_id", "").strip()

    if not student_no:
        raise ValidationError("学号不能为空。")
    if len(student_no) > 20:
        raise ValidationError("学号不能超过 20 个字符。")
    if not name:
        raise ValidationError("姓名不能为空。")
    if len(name) > 50:
        raise ValidationError("姓名不能超过 50 个字符。")
    if gender not in PLAYER_GENDER_OPTIONS:
        raise ValidationError("请选择有效的性别。")
    if level_code not in PLAYER_LEVEL_LABELS:
        raise ValidationError("请选择有效的运动等级。")
    if injury_status_code not in PLAYER_INJURY_STATUS_LABELS:
        raise ValidationError("请选择有效的伤病状态。")
    if len(play_style) > 100:
        raise ValidationError("主打法不能超过 100 个字符。")
    if len(grip) > 50:
        raise ValidationError("握拍方式不能超过 50 个字符。")
    if len(contact_phone) > 20:
        raise ValidationError("联系电话不能超过 20 个字符。")

    if not coach_id_text.isdigit():
        raise ValidationError("请选择有效的所属教练。")

    coach = next((item for item in COACHES if item["id"] == int(coach_id_text)), None)
    if not coach:
        raise ValidationError("所属教练不存在。")

    level = PLAYER_LEVEL_LABELS[level_code]
    injury_status = PLAYER_INJURY_STATUS_LABELS[injury_status_code]
    return {
        "student_no": student_no,
        "name": name,
        "gender": gender,
        "age": age,
        "level": level,
        "skill_level": level,
        "level_code": level_code,
        "play_style": play_style,
        "grip": grip,
        "contact_phone": contact_phone,
        "injury_status": injury_status,
        "injury_status_code": injury_status_code,
        "coach_id": coach["id"],
        "coach_name": coach["name"],
    }


def find_player(athlete_id):
    return next((item for item in PLAYERS if item["id"] == athlete_id), None)


def filter_rehab_records(args):
    predicates = []
    player_keyword = args.get("player_keyword", "").strip().lower()
    risk_level = args.get("risk_level", "").strip()
    recovery_status = args.get("recovery_status", "").strip()
    overdue_only = args.get("overdue_only", "").strip()

    records = [
        enrich_rehab_record(item)
        for item in INJURY_RECORDS
        if not item.get("is_deleted") and item["recovery_status"] != "已恢复"
    ]

    if player_keyword:
        predicates.append(
            lambda record, value=player_keyword: value in record["player_name"].lower()
            or value in record["student_no"].lower()
            or value in record["injury_location"].lower()
        )
    if risk_level:
        predicates.append(lambda record, value=risk_level: record["risk_code"] == value)
    if recovery_status:
        predicates.append(lambda record, value=recovery_status: record["recovery_status"] == value)
    if overdue_only:
        predicates.append(lambda record: record["is_overdue"])

    records.sort(key=lambda item: (item["risk_weight"], item["expected_recovery_date"], item["id"]), reverse=True)
    if not predicates:
        return records, 0
    return [record for record in records if all(check(record) for check in predicates)], len(predicates)


def enrich_rehab_record(record):
    player = find_player(record["athlete_id"])
    followup = latest_followup(record["id"])
    pain_score = followup["pain_score"] if followup else estimate_pain_score(record)
    progress = calculate_rehab_progress(record)
    is_overdue = bool(
        record.get("expected_recovery_date")
        and record["expected_recovery_date"] < datetime.now().strftime("%Y-%m-%d")
    )
    risk = evaluate_rehab_risk(record, pain_score, progress, is_overdue)
    base = enrich_injury_record(record)
    base.update(
        {
            "player_name": player["name"] if player else "未知运动员",
            "student_no": player["student_no"] if player else "-",
            "pain_score": pain_score,
            "progress": progress,
            "risk_code": risk["code"],
            "risk_label": risk["label"],
            "risk_class": risk["class"],
            "risk_weight": risk["weight"],
            "latest_followup_date": followup["followup_date"] if followup else "-",
            "training_limit": followup["training_limit"] if followup else build_training_alert_for_record(record),
            "reviewer": followup["reviewer"] if followup else "待复诊",
            "return_training_advice": build_return_training_advice(record, pain_score, progress, risk["code"]),
        }
    )
    return base


def estimate_pain_score(record):
    if record["severity"] == "严重":
        return 6
    if record["severity"] == "中度":
        return 4
    return 2


def calculate_rehab_progress(record):
    injury_date = parse_date_for_calc(record["injury_date"])
    expected_date = parse_date_for_calc(record.get("expected_recovery_date", ""))
    if not injury_date or not expected_date:
        return 35
    total_days = max((expected_date - injury_date).days, 1)
    elapsed_days = max((datetime.now() - injury_date).days, 0)
    return max(5, min(100, round(elapsed_days / total_days * 100)))


def evaluate_rehab_risk(record, pain_score, progress, is_overdue):
    if record["severity"] == "严重" and record["recovery_status"] == "治疗中":
        return {"code": "high", "label": "高风险", "class": "danger", "weight": 3}
    if is_overdue or pain_score >= 6 or progress < 35:
        return {"code": "high", "label": "高风险", "class": "danger", "weight": 3}
    if record["severity"] == "严重" or pain_score >= 3 or record["recovery_status"] == "康复中":
        return {"code": "medium", "label": "观察", "class": "warning", "weight": 2}
    return {"code": "low", "label": "可控", "class": "success", "weight": 1}


def build_return_training_advice(record, pain_score, progress, risk_code):
    if risk_code == "high":
        return "暂停对抗和高强度专项训练，先完成复诊评估。"
    if record["recovery_status"] == "康复中" and pain_score <= 3 and progress >= 70:
        return "可安排低到中等强度复训，逐步恢复步法和多球训练。"
    if record["recovery_status"] == "康复中":
        return "维持康复训练，控制单次负荷并保留疼痛反馈。"
    return "以治疗和基础活动度恢复为主，暂不进入正常训练。"


def build_rehab_summary(records):
    risk_counts = {"high": 0, "medium": 0, "low": 0}
    progress_total = 0
    pain_total = 0
    ready_count = 0
    overdue_count = 0
    for record in records:
        risk_counts[record["risk_code"]] += 1
        progress_total += record["progress"]
        pain_total += record["pain_score"]
        if record["risk_code"] == "low" and record["progress"] >= 70:
            ready_count += 1
        if record["is_overdue"]:
            overdue_count += 1
    total = len(records)
    return {
        "record_count": total,
        "high_risk_count": risk_counts["high"],
        "ready_count": ready_count,
        "overdue_count": overdue_count,
        "average_progress": round(progress_total / total, 1) if total else 0,
        "average_pain": round(pain_total / total, 1) if total else 0,
        "risk_pie": [
            {"name": "高风险", "value": risk_counts["high"]},
            {"name": "观察", "value": risk_counts["medium"]},
            {"name": "可控", "value": risk_counts["low"]},
        ],
    }


def filter_match_records(args):
    predicates = []
    player_keyword = args.get("player_keyword", "").strip().lower()
    result = args.get("result", "").strip()
    date_from = args.get("date_from", "").strip()
    date_to = args.get("date_to", "").strip()
    opponent_keyword = args.get("opponent_keyword", "").strip().lower()

    if player_keyword:
        predicates.append(
            lambda record, value=player_keyword: value in record["player_name"].lower()
            or value in record["student_no"].lower()
        )
    if result:
        predicates.append(lambda record, value=result: record["result"] == value)
    if date_from:
        predicates.append(lambda record, value=date_from: record["match_date"] >= value)
    if date_to:
        predicates.append(lambda record, value=date_to: record["match_date"] <= value)
    if opponent_keyword:
        predicates.append(lambda record, value=opponent_keyword: value in record["opponent"].lower())

    records = [enrich_match_record(item) for item in MATCH_RESULTS]
    records.sort(key=lambda item: (item["match_date"], item["id"]), reverse=True)
    if not predicates:
        return records, 0
    return [record for record in records if all(check(record) for check in predicates)], len(predicates)


def enrich_match_record(record):
    player = find_player(record["athlete_id"])
    opponent_player = find_player(record.get("opponent_id"))
    if not opponent_player and record.get("opponent"):
        opponent_player = next((item for item in PLAYERS if item["name"] == record["opponent"]), None)
    opponent_source = record.get("opponent_source")
    if opponent_source not in {"internal", "external"}:
        opponent_source = "internal" if opponent_player else "external"
    match_type = record.get("match_type") or "正式比赛"
    base = dict(record)
    base.update(
        {
            "player_name": player["name"] if player else "未知运动员",
            "student_no": player["student_no"] if player else "-",
            "level": (player.get("skill_level") or player.get("level")) if player else "-",
            "match_type": match_type,
            "opponent_source": opponent_source,
            "opponent_id": opponent_player["id"] if opponent_source == "internal" and opponent_player else record.get("opponent_id"),
            "opponent": opponent_player["name"] if opponent_source == "internal" and opponent_player else record.get("opponent", ""),
            "opponent_external_name": record.get("opponent", "") if opponent_source == "external" else "",
            "opponent_student_no": opponent_player["student_no"] if opponent_source == "internal" and opponent_player else "",
            "opponent_source_label": "系统内" if opponent_source == "internal" else "系统外",
            "result_class": match_result_class(record["result"]),
            "score_diff": calculate_match_score_diff(record["score"]),
        }
    )
    return base


def get_editing_match_record(edit_id):
    if not edit_id.isdigit():
        return None
    record = next((item for item in MATCH_RESULTS if item["id"] == int(edit_id)), None)
    return enrich_match_record(record) if record else None


def save_match_record(form):
    validated = validate_match_form(form)
    record_id = validated.pop("record_id")
    if record_id is None:
        MATCH_RESULTS.append({"id": next_id(MATCH_RESULTS), **validated})
        return

    target = next((item for item in MATCH_RESULTS if item["id"] == record_id), None)
    if not target:
        raise ValidationError("要修改的比赛成绩不存在。")
    target.update(validated)


def get_match_record(match_id):
    return next((item for item in MATCH_RESULTS if item["id"] == match_id), None)


def next_match_analysis_version(match_id):
    versions = [
        analysis["version_no"]
        for analysis in MATCH_TACTICAL_ANALYSES
        if analysis["match_id"] == match_id
    ]
    return max(versions, default=0) + 1


def get_match_analysis(analysis_id):
    return next((item for item in MATCH_TACTICAL_ANALYSES if item["id"] == analysis_id), None)


def get_latest_confirmed_analysis_for_match(match_id):
    confirmed = [
        analysis
        for analysis in MATCH_TACTICAL_ANALYSES
        if analysis["match_id"] == match_id and analysis["status"] == "confirmed"
    ]
    return max(confirmed, key=lambda item: (item["version_no"], item["id"]), default=None)


def save_match_tactical_analysis(match_id, form, operator):
    match = get_match_record(match_id)
    if not match:
        raise ValidationError("比赛记录不存在。")
    analysis_method = form.get("analysis_method", "").strip() or "three_phase"
    if analysis_method != "three_phase":
        raise ValidationError("当前仅开放三段法录入框架。")
    status = form.get("status", "").strip() or "draft"
    if status not in {"draft", "confirmed"}:
        raise ValidationError("分析状态只能为草稿或已确认。")

    phase_inputs = []
    for phase_code in THREE_PHASE_LABELS:
        phase_inputs.append(
            {
                "phase_code": phase_code,
                "points_won": form.get(f"{phase_code}_points_won", "0"),
                "points_lost": form.get(f"{phase_code}_points_lost", "0"),
            }
        )
    phase_stats = calculate_three_phase_stats(phase_inputs, standard_verified=False)

    global MATCH_ANALYSIS_ID_COUNTER, MATCH_PHASE_STAT_ID_COUNTER
    analysis_id = MATCH_ANALYSIS_ID_COUNTER
    MATCH_ANALYSIS_ID_COUNTER += 1
    analysis = {
        "id": analysis_id,
        "match_id": match_id,
        "analysis_method": analysis_method,
        "version_no": next_match_analysis_version(match_id),
        "status": status,
        "coach_summary": form.get("coach_summary", "").strip(),
        "standard_verification_status": "pending",
        "created_by": operator,
    }
    MATCH_TACTICAL_ANALYSES.append(analysis)
    for stat in phase_stats:
        MATCH_PHASE_STATS.append(
            {
                "id": MATCH_PHASE_STAT_ID_COUNTER,
                "analysis_id": analysis_id,
                **stat,
            }
        )
        MATCH_PHASE_STAT_ID_COUNTER += 1
    return analysis


def build_match_analysis_view_models():
    views = []
    for analysis in MATCH_TACTICAL_ANALYSES:
        match = get_match_record(analysis["match_id"])
        stats = [
            stat
            for stat in MATCH_PHASE_STATS
            if stat["analysis_id"] == analysis["id"]
        ]
        views.append(
            {
                **analysis,
                "match_name": match["match_name"] if match else "未知比赛",
                "athlete_name": find_player(match["athlete_id"])["name"] if match and find_player(match["athlete_id"]) else "未知运动员",
                "status_label": "已确认" if analysis["status"] == "confirmed" else "草稿",
                "stats": stats,
            }
        )
    return sorted(views, key=lambda item: item["id"], reverse=True)


def build_match_plan_source_summary(match):
    return f"{match['match_name']} · {match['result']} {match['score']}"


def build_match_analysis_plan_source_summary(analysis):
    match = get_match_record(analysis["match_id"])
    match_summary = build_match_plan_source_summary(match) if match else "未知比赛"
    coach_summary = analysis.get("coach_summary") or "教练未填写结论"
    return f"{match_summary} · 已确认技战术分析 V{analysis['version_no']} · 标准待核验 · {coach_summary}"


THREE_PHASE_LABELS = {
    "serve_attack": "发抢段",
    "receive_attack": "接抢段",
    "rally": "相持段",
}


def calculate_three_phase_stats(phase_inputs, standard_verified=False):
    normalized = []
    total_points = 0
    for item in phase_inputs:
        phase_code = item["phase_code"]
        try:
            points_won = int(item["points_won"])
            points_lost = int(item["points_lost"])
        except (TypeError, ValueError):
            raise ValidationError("三段法得分和失分必须为非负整数。")
        if points_won < 0 or points_lost < 0:
            raise ValidationError("三段法得分和失分必须为非负整数。")
        phase_total = points_won + points_lost
        total_points += phase_total
        normalized.append(
            {
                "phase_code": phase_code,
                "phase_label": THREE_PHASE_LABELS.get(phase_code, phase_code),
                "points_won": points_won,
                "points_lost": points_lost,
                "phase_total": phase_total,
            }
        )

    result = []
    for item in normalized:
        phase_total = item["phase_total"]
        scoring_rate = round(item["points_won"] / phase_total * 100, 2) if phase_total else None
        usage_rate = round(phase_total / total_points * 100, 2) if phase_total and total_points else None
        enriched = dict(item)
        enriched.update(
            {
                "scoring_rate": scoring_rate,
                "usage_rate": usage_rate,
                "usage_denominator_source": "三段得失分合计" if total_points else "",
                "evaluation_level": None if not standard_verified else None,
                "display_note": "无该段数据" if phase_total == 0 else "",
            }
        )
        result.append(enriched)
    return result


TRAINING_PLAN_ITEM_MODULE_LABELS = {
    "fitness": "体能训练",
    "footwork": "步法训练",
    "technique_tactic": "技战术训练",
}


def parse_training_plan_items(form):
    module_types = form.getlist("item_module_type")
    titles = form.getlist("item_title")
    targets = form.getlist("item_target_description")
    sessions = form.getlist("item_planned_sessions")
    minutes = form.getlist("item_planned_minutes")
    intensities = form.getlist("item_intensity")
    max_len = max(
        len(module_types),
        len(titles),
        len(targets),
        len(sessions),
        len(minutes),
        len(intensities),
        0,
    )
    items = []
    errors = []

    for index in range(max_len):
        module_type = module_types[index].strip() if index < len(module_types) else ""
        title = titles[index].strip() if index < len(titles) else ""
        target = targets[index].strip() if index < len(targets) else ""
        planned_sessions = sessions[index].strip() if index < len(sessions) else ""
        planned_minutes = minutes[index].strip() if index < len(minutes) else ""
        intensity = intensities[index].strip() if index < len(intensities) else ""

        if not any([module_type, title, target, planned_sessions, planned_minutes, intensity]):
            continue
        row_label = f"第{index + 1}个计划项目"
        if module_type not in TRAINING_PLAN_ITEM_MODULE_LABELS:
            errors.append(f"{row_label}模块类型必须为体能训练、步法训练或技战术训练。")
        if not title:
            errors.append(f"{row_label}标题不能为空。")
        if not target:
            errors.append(f"{row_label}目标不能为空。")
        try:
            planned_sessions_int = int(planned_sessions)
            if planned_sessions_int <= 0:
                raise ValueError
        except ValueError:
            errors.append(f"{row_label}计划次数必须为正整数。")
            planned_sessions_int = 0
        try:
            planned_minutes_int = int(planned_minutes)
            if planned_minutes_int <= 0:
                raise ValueError
        except ValueError:
            errors.append(f"{row_label}计划时长必须为正整数。")
            planned_minutes_int = 0
        if intensity not in {"低", "中", "高", "极高"}:
            errors.append(f"{row_label}强度必须为低、中、高或极高。")

        items.append(
            {
                "module_type": module_type,
                "module_label": TRAINING_PLAN_ITEM_MODULE_LABELS.get(module_type, module_type),
                "item_title": title,
                "target_description": target,
                "planned_sessions": planned_sessions_int,
                "planned_minutes": planned_minutes_int,
                "intensity": intensity,
                "status": "pending",
            }
        )

    if not items:
        errors.append("至少添加一个训练计划项目。")
    return items, errors


def group_training_plan_items_by_plan():
    grouped = {}
    for item in TRAINING_PLAN_ITEMS:
        grouped.setdefault(item["plan_id"], []).append(item)
    return grouped


def get_training_plan_items(plan_id):
    return [item for item in TRAINING_PLAN_ITEMS if item["plan_id"] == plan_id]


def validate_training_plan_status_choice(plan_status, plan_items):
    errors = []
    if plan_status in EXECUTABLE_PLAN_STATUSES and not plan_items:
        errors.append("没有计划项目不能发布训练计划。")
    if plan_status == "completed" and any(item.get("status") != "completed" for item in plan_items):
        errors.append("所有计划项目完成后才能标记训练计划已完成。")
    return errors


def is_plan_released_for_execution(plan):
    if not plan:
        return False
    return plan.get("status", "published") in EXECUTABLE_PLAN_STATUSES


def get_training_plan_status_label(plan):
    return TRAINING_PLAN_STATUS_LABELS.get(plan.get("status", "published"), "已发布")


def build_execution_frame_items(module_type):
    items = []
    for item in TRAINING_PLAN_ITEMS:
        if item["module_type"] != module_type:
            continue
        plan = next((plan for plan in TRAINING_PLANS if plan["id"] == item["plan_id"]), None)
        if not is_plan_released_for_execution(plan):
            continue
        athlete = find_player(plan["athlete_id"]) if plan else None
        enriched = dict(item)
        enriched["title"] = item["item_title"]
        enriched["plan_title"] = plan["content"] if plan else ""
        enriched["athlete_name"] = athlete["name"] if athlete else "未知运动员"
        items.append(enriched)
    return items


def validate_match_form(form):
    record_id = form.get("record_id", "").strip()
    athlete_id = parse_int_field(form.get("athlete_id", "").strip(), "运动员")
    if not find_player(athlete_id):
        raise ValidationError("所选运动员不存在，请重新选择。")

    match_date = parse_date_field(form.get("match_date", "").strip(), "比赛日期")
    match_name = form.get("match_name", "").strip()
    if not match_name:
        raise ValidationError("比赛名称不能为空。")

    match_type = form.get("match_type", "").strip() or "正式比赛"
    if match_type not in {"友谊赛", "正式比赛"}:
        raise ValidationError("比赛类型非法，请选择友谊赛或正式比赛。")

    opponent_source = form.get("opponent_source", "").strip() or "internal"
    if opponent_source not in {"internal", "external"}:
        raise ValidationError("对手来源非法，请选择系统内运动员或系统外参赛者。")

    opponent_id = None
    opponent_name = ""
    if opponent_source == "internal":
        opponent_id = parse_int_field(form.get("opponent_id", "").strip(), "对手")
        opponent_player = find_player(opponent_id)
        if not opponent_player:
            raise ValidationError("所选对手不存在，请重新选择。")
        if opponent_id == athlete_id:
            raise ValidationError("对手不能与运动员相同。")
        opponent_name = opponent_player["name"]
    else:
        opponent_name = form.get("opponent_external_name", "").strip()
        if not opponent_name:
            raise ValidationError("系统外对手姓名不能为空。")

    result = form.get("result", "").strip()
    if result not in {"胜", "负", "平"}:
        raise ValidationError("比赛结果非法，请从胜、负、平中选择。")

    score = form.get("score", "").strip()
    if not score:
        raise ValidationError("比分不能为空。")

    return {
        "record_id": int(record_id) if record_id.isdigit() else None,
        "athlete_id": athlete_id,
        "match_type": match_type,
        "opponent_source": opponent_source,
        "match_date": match_date,
        "match_name": match_name,
        "opponent_id": opponent_id,
        "opponent": opponent_name,
        "result": result,
        "score": score,
        "rank": form.get("rank", "").strip(),
        "key_points": form.get("key_points", "").strip(),
        "tactic_review": form.get("tactic_review", "").strip(),
        "improvement": form.get("improvement", "").strip(),
    }


def match_result_class(result):
    return {
        "胜": "success",
        "负": "danger",
        "平": "secondary",
    }.get(result, "secondary")


def calculate_match_score_diff(score):
    parts = score.split(":")
    if len(parts) != 2:
        return 0
    try:
        return int(parts[0]) - int(parts[1])
    except ValueError:
        return 0


def build_match_summary(records):
    result_counts = {"胜": 0, "负": 0, "平": 0}
    month_map = {}
    player_map = {}
    for record in records:
        result_counts[record["result"]] += 1
        month_key = record["match_date"][:7]
        month_map[month_key] = month_map.get(month_key, 0) + 1
        player_stats = player_map.setdefault(record["player_name"], {"matches": 0, "wins": 0})
        player_stats["matches"] += 1
        if record["result"] == "胜":
            player_stats["wins"] += 1

    total = len(records)
    player_rates = sorted(
        (
            {
                "name": name,
                "matches": stats["matches"],
                "win_rate": round(stats["wins"] / stats["matches"] * 100, 1) if stats["matches"] else 0,
            }
            for name, stats in player_map.items()
        ),
        key=lambda item: (item["win_rate"], item["matches"]),
        reverse=True,
    )
    month_labels = sorted(month_map.keys())
    return {
        "record_count": total,
        "win_count": result_counts["胜"],
        "loss_count": result_counts["负"],
        "draw_count": result_counts["平"],
        "win_rate": round(result_counts["胜"] / total * 100, 1) if total else 0,
        "result_pie": [{"name": key, "value": value} for key, value in result_counts.items()],
        "month_labels": month_labels,
        "month_counts": [month_map[key] for key in month_labels],
        "player_rates": player_rates[:5],
    }


def build_user_permission_summary(users):
    admin_count = sum(1 for user in users if user["role"] == "admin")
    coach_count = sum(1 for user in users if user["role"] == "coach")
    return {
        "total_count": len(users),
        "admin_count": admin_count,
        "coach_count": coach_count,
        "database_account_count": len(DATABASE_ACCOUNTS),
    }


def build_import_templates():
    return [
        {"name": "训练计划导入模板", "fields": ["运动员", "教练", "训练日期", "内容", "强度", "时长", "地点"]},
        {"name": "步法训练导入模板", "fields": ["运动员学号", "训练日期", "步法类型", "训练时长(分钟)", "训练组数", "训练备注"]},
        {"name": "体能测试导入模板", "fields": ["运动员", "测试日期", "测试教练", "上肢力量", "下肢力量", "柔韧性", "耐力", "速度"]},
        {"name": "比赛成绩导入模板", "fields": ["运动员", "比赛日期", "比赛名称", "对手", "结果", "比分", "复盘备注"]},
    ]


def parse_date_for_calc(value):
    if not value:
        return None
    try:
        return datetime.strptime(value[:10], "%Y-%m-%d")
    except ValueError:
        return None


class ValidationError(Exception):
    pass


def parse_percent_field(value, field_name, required=False):
    text = str(value or "").strip()
    if not text:
        if required:
            raise ValidationError(f"{field_name}不能为空。")
        return None
    try:
        number = float(text)
    except ValueError:
        raise ValidationError(f"{field_name}必须是 0 到 100 之间的数字。")
    if not 0 <= number <= 100:
        raise ValidationError(f"{field_name}必须在 0 到 100 之间。")
    return round(number, 2)


def parse_landing_distribution_items(form):
    items = []
    for landing in get_landing_options():
        landing_id = landing["id"]
        concentration = str(form.get(f"landing_concentration_{landing_id}", "")).strip()
        if not concentration:
            continue
        if concentration not in LANDING_CONCENTRATION_OPTIONS:
            raise ValidationError(f"落点「{landing['dict_name']}」的集中程度选项非法。")
        items.append({"landing_dict_id": landing_id, "concentration": concentration})
    return items


def build_landing_distribution_map(items):
    return {item["landing_dict_id"]: item["concentration"] for item in items or []}


def enrich_footwork_record(record):
    player = find_player(record["athlete_id"])
    footwork = find_dict_by_id(record["footwork_dict_id"])
    base = dict(record)
    base.update(
        {
            "athlete_name": player["name"] if player else record.get("athlete_name", "未知运动员"),
            "student_no": player["student_no"] if player else "-",
            "level": (player.get("skill_level") or player.get("level")) if player else "-",
            "footwork_type_name": footwork["dict_name"] if footwork else "-",
        }
    )
    return base


def enrich_technique_tactic_record(record):
    player = find_player(record["athlete_id"])
    technique = find_dict_by_id(record["technique_dict_id"])
    category = find_dict_by_id(technique["parent_id"]) if technique else None
    base = dict(record)
    base.update(
        {
            "athlete_name": player["name"] if player else record.get("athlete_name", "未知运动员"),
            "student_no": player["student_no"] if player else "-",
            "level": (player.get("skill_level") or player.get("level")) if player else "-",
            "technique_name": technique["dict_name"] if technique else "-",
            "technique_category_name": category["dict_name"] if category else "-",
            "technique_category_id": category["id"] if category else None,
            "landing_distribution": record.get("landing_distribution") or "-",
            "landing_distribution_map": build_landing_distribution_map(record.get("landing_distribution_items", [])),
        }
    )
    return base


def get_editing_footwork_record(edit_id):
    if not edit_id.isdigit():
        return None
    record = training_repo.get_footwork_record(int(edit_id))
    return enrich_footwork_record(record) if record else None


def get_editing_technique_tactic_record(edit_id):
    if not edit_id.isdigit():
        return None
    record = training_repo.get_technique_tactic_record(int(edit_id))
    return enrich_technique_tactic_record(record) if record else None


def save_footwork_training_record(form, operator, record_id=None):
    validated = validate_footwork_training_form(form)
    if record_id is None:
        training_repo.create_footwork_record(validated, operator)
        return

    if not training_repo.update_footwork_record(record_id, validated, operator):
        raise ValidationError("要修改的步法训练记录不存在。")


def save_technique_tactic_record(form, operator, record_id=None):
    validated = validate_technique_tactic_form(form)
    if record_id is None:
        training_repo.create_technique_tactic_record(validated, operator)
        return

    if not training_repo.update_technique_tactic_record(record_id, validated, operator):
        raise ValidationError("要修改的技战术训练记录不存在。")


def validate_footwork_training_form(form):
    athlete_id = parse_int_field(str(form.get("athlete_id", "")).strip(), "运动员")
    player = find_player(athlete_id)
    if not player:
        raise ValidationError("所选运动员不存在，请重新选择。")

    training_date = parse_date_field(str(form.get("training_date", "")).strip(), "训练日期")
    footwork_dict_id = parse_int_field(str(form.get("footwork_dict_id", "")).strip(), "步法类型")
    footwork = find_dict_by_id(footwork_dict_id)
    if not footwork or footwork["category_type"] != "footwork" or footwork["dict_code"] == "footwork_root":
        raise ValidationError("步法类型非法，请从页面选项中选择。")

    duration_minutes = parse_int_range(str(form.get("duration_minutes", "")).strip(), "训练时长", 1, 300)
    set_count = parse_int_range(str(form.get("set_count", "")).strip(), "训练组数", 1, 100)
    note = str(form.get("training_note", form.get("note", ""))).strip()
    if len(note) > 500:
        raise ValidationError("训练备注不能超过 500 个字符。")

    return {
        "athlete_id": athlete_id,
        "athlete_name": player["name"],
        "training_date": training_date,
        "footwork_dict_id": footwork_dict_id,
        "duration_minutes": duration_minutes,
        "set_count": set_count,
        "note": note,
    }


def validate_technique_tactic_form(form):
    athlete_id = parse_int_field(str(form.get("athlete_id", "")).strip(), "运动员")
    player = find_player(athlete_id)
    if not player:
        raise ValidationError("所选运动员不存在，请重新选择。")

    training_date = parse_date_field(str(form.get("training_date", "")).strip(), "训练日期")
    technique_dict_id = parse_int_field(str(form.get("technique_dict_id", "")).strip(), "技战术")
    technique = find_dict_by_id(technique_dict_id)
    if not technique or technique["category_type"] != "technique_tactic" or technique["parent_id"] == 0:
        raise ValidationError("技战术选择非法，请选择二级技战术项目。")

    multi_ball_count = parse_int_range(str(form.get("multi_ball_count", "")).strip(), "多球训练球数", 0, 5000)
    serve_frequency = str(form.get("serve_frequency", "")).strip()
    if serve_frequency not in SERVE_FREQUENCY_OPTIONS:
        raise ValidationError("发球频率非法，请从页面选项中选择。")

    plan_execution_rate = parse_percent_field(form.get("plan_execution_rate"), "计划执行率", required=True)
    on_table_rate = parse_percent_field(form.get("on_table_rate"), "上台率", required=False)
    landing_distribution_items = parse_landing_distribution_items(form)
    qualitative_comment = str(form.get("qualitative_comment", "")).strip()
    if len(qualitative_comment) > 500:
        raise ValidationError("定性描述不能超过 500 个字符。")

    return {
        "athlete_id": athlete_id,
        "athlete_name": player["name"],
        "training_date": training_date,
        "technique_dict_id": technique_dict_id,
        "multi_ball_count": multi_ball_count,
        "serve_frequency": serve_frequency,
        "plan_execution_rate": plan_execution_rate,
        "on_table_rate": on_table_rate,
        "landing_distribution_items": landing_distribution_items,
        "qualitative_comment": qualitative_comment,
    }


def _training_filter_args(args):
    if hasattr(args, "get"):
        return {
            "athlete_id": args.get("athlete_id", "").strip(),
            "start_date": args.get("start_date", "").strip(),
            "end_date": args.get("end_date", "").strip(),
            "footwork_dict_id": args.get("footwork_dict_id", "").strip(),
            "technique_category_id": args.get("technique_category_id", "").strip(),
            "technique_dict_id": args.get("technique_dict_id", "").strip(),
            "serve_frequency": args.get("serve_frequency", "").strip(),
            "keyword": args.get("keyword", "").strip(),
        }
    return args


def _count_active_training_filters(filters, keys):
    return sum(1 for key in keys if filters.get(key))


def filter_footwork_training_records(args):
    filters = _training_filter_args(args)
    records = [enrich_footwork_record(item) for item in training_repo.list_footwork_records(filters)]
    active_keys = ("athlete_id", "start_date", "end_date", "footwork_dict_id", "keyword")
    return records, _count_active_training_filters(filters, active_keys)


def filter_technique_tactic_records(args):
    filters = _training_filter_args(args)
    records = [enrich_technique_tactic_record(item) for item in training_repo.list_technique_tactic_records(filters)]
    active_keys = ("athlete_id", "start_date", "end_date", "technique_category_id", "technique_dict_id", "serve_frequency", "keyword")
    return records, _count_active_training_filters(filters, active_keys)


def build_overall_stats():
    monthly_training = {}
    for plan in TRAINING_PLANS:
        month_key = plan["plan_datetime"][:7]
        stats = monthly_training.setdefault(month_key, {"duration": 0, "count": 0})
        stats["duration"] += plan["duration"]
        stats["count"] += 1

    month_labels = sorted(monthly_training)
    monthly_duration = [monthly_training[key]["duration"] for key in month_labels]
    monthly_train_count = [monthly_training[key]["count"] for key in month_labels]

    injury_locations = {}
    for record in INJURY_RECORDS:
        if record.get("is_deleted"):
            continue
        location = record["injury_location"]
        injury_locations[location] = injury_locations.get(location, 0) + 1
    injury_pie = [{"name": name, "value": value} for name, value in sorted(injury_locations.items())]

    latest_fitness = {}
    for test in sorted(FITNESS_TESTS, key=lambda item: item["test_date"]):
        latest_fitness[test["athlete_id"]] = test
    fitness_rank = []
    for athlete_id, test in latest_fitness.items():
        player = find_player(athlete_id)
        if player:
            fitness_rank.append({"name": player["name"], "score": test["overall_score"]})
    fitness_rank.sort(key=lambda item: item["score"], reverse=True)

    intensity_counts = {}
    for plan in TRAINING_PLANS:
        intensity = plan["intensity"]
        intensity_counts[intensity] = intensity_counts.get(intensity, 0) + 1
    intensity_pie = [{"name": name, "value": value} for name, value in sorted(intensity_counts.items())]

    monthly_skill = {}
    for record in training_repo.list_technique_tactic_records():
        month_key = record["training_date"][:7]
        score = calculate_technique_tactic_score(record)
        stats = monthly_skill.setdefault(month_key, {"total": 0.0, "count": 0})
        stats["total"] += score
        stats["count"] += 1
    skill_month_labels = sorted(monthly_skill)
    skill_month_scores = [
        round(monthly_skill[key]["total"] / monthly_skill[key]["count"], 1)
        for key in skill_month_labels
    ]

    current_month = datetime.now().strftime("%Y-%m")
    active_injury_statuses = {"治疗中", "康复中"}
    active_injuries = sum(
        1
        for record in INJURY_RECORDS
        if not record.get("is_deleted") and record["recovery_status"] in active_injury_statuses
    )
    fitness_scores = [item["score"] for item in fitness_rank]

    return {
        "cards": {
            "total_athletes": len(PLAYERS),
            "current_month_duration": monthly_training.get(current_month, {"duration": 0})["duration"],
            "active_injuries": active_injuries,
            "avg_fitness": round(sum(fitness_scores) / len(fitness_scores), 1) if fitness_scores else 0,
        },
        "month_labels": month_labels,
        "monthly_duration": monthly_duration,
        "monthly_train_count": monthly_train_count,
        "injury_pie": injury_pie,
        "fitness_player_names": [item["name"] for item in fitness_rank],
        "fitness_player_scores": fitness_scores,
        "intensity_pie": intensity_pie,
        "skill_month_labels": skill_month_labels,
        "skill_month_scores": skill_month_scores,
    }


def build_home_dashboard_data():
    stats = build_overall_stats()
    intensity_order = ["低", "中", "高", "极高"]
    intensity_month_map = {}
    for plan in TRAINING_PLANS:
        month_key = plan["plan_datetime"][:7]
        month_stats = intensity_month_map.setdefault(
            month_key,
            {intensity: 0 for intensity in intensity_order},
        )
        month_stats[plan["intensity"]] = month_stats.get(plan["intensity"], 0) + plan["duration"]

    intensity_month_labels = sorted(intensity_month_map)
    intensity_series = [
        {
            "name": f"{intensity}强度",
            "data": [intensity_month_map[month].get(intensity, 0) for month in intensity_month_labels],
        }
        for intensity in intensity_order
    ]

    fitness_radar_indicators = [
        {"name": "上肢力量", "max": 100},
        {"name": "下肢力量", "max": 100},
        {"name": "柔韧性", "max": 100},
        {"name": "耐力", "max": 100},
        {"name": "速度", "max": 100},
        {"name": "爆发力", "max": 100},
        {"name": "敏捷", "max": 100},
        {"name": "核心稳定", "max": 100},
        {"name": "移动效率", "max": 100},
        {"name": "恢复指数", "max": 100},
    ]
    if FITNESS_TESTS:
        base_metrics = [
            round(sum(test["upper_strength"] for test in FITNESS_TESTS) / len(FITNESS_TESTS), 1),
            round(sum(test["lower_strength"] for test in FITNESS_TESTS) / len(FITNESS_TESTS), 1),
            round(sum(test["flexibility"] for test in FITNESS_TESTS) / len(FITNESS_TESTS), 1),
            round(sum(test["endurance"] for test in FITNESS_TESTS) / len(FITNESS_TESTS), 1),
            round(sum(test["speed"] for test in FITNESS_TESTS) / len(FITNESS_TESTS), 1),
        ]
        upper_strength, lower_strength, flexibility, endurance, speed = base_metrics
        derived_metrics = [
            round((lower_strength + speed) / 2, 1),
            round((speed + flexibility) / 2, 1),
            round((upper_strength + lower_strength + flexibility) / 3, 1),
            round((speed + endurance) / 2, 1),
            round((flexibility + endurance) / 2, 1),
        ]
        fitness_radar_values = base_metrics + derived_metrics
    else:
        fitness_radar_values = [0 for _ in fitness_radar_indicators]

    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "cards": {
            **stats["cards"],
            "training_plan_count": len(TRAINING_PLANS),
        },
        "month_labels": stats["month_labels"],
        "monthly_duration": stats["monthly_duration"],
        "injury_pie": stats["injury_pie"],
        "intensity_month_labels": intensity_month_labels,
        "intensity_series": intensity_series,
        "fitness_radar_indicators": fitness_radar_indicators,
        "fitness_radar_values": fitness_radar_values,
        "fitness_target_values": [85, 86, 82, 86, 88, 87, 84, 85, 86, 83],
    }


def calculate_technique_tactic_score(record):
    if record.get("on_table_rate") is not None:
        return float(record["on_table_rate"])
    return float(record.get("plan_execution_rate") or 0)


def footwork_record_identity(record):
    normalize_text = lambda value: " ".join(str(value or "").split())
    return (
        record["athlete_id"],
        record["training_date"],
        record["footwork_dict_id"],
        record["duration_minutes"],
        record["set_count"],
        normalize_text(record.get("note")),
    )


def import_footwork_training_excel(file, operator):
    imported_count = 0
    error_rows = []
    existing_record_identities = set()

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [str(cell.value or "").strip() for cell in ws[1]]
        excel_format = detect_footwork_excel_format(headers)
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if excel_format == "legacy_skill":
                values = (list(row) + [None] * 7)[:7]
                if not any(values):
                    continue
                athlete_value, training_date, footwork_value, _, duration_value, _, note_value = values
                form = build_legacy_footwork_import_form(
                    athlete_value,
                    training_date,
                    footwork_value,
                    duration_value,
                    note_value,
                )
            elif excel_format == "member9":
                values = (list(row) + [None] * 6)[:6]
                if not any(values):
                    continue
                athlete_value, training_date, footwork_duration, hit_score, multi_ball_duration, intensity_value = values
                form = build_member9_footwork_import_form(
                    athlete_value,
                    training_date,
                    footwork_duration,
                    hit_score,
                    multi_ball_duration,
                    intensity_value,
                )
            else:
                values = (list(row) + [None] * 6)[:6]
                if not any(values):
                    continue
                student_no, training_date, footwork_value, duration_value, set_value, note_value = values
                form = build_footwork_import_form(
                    student_no,
                    training_date,
                    footwork_value,
                    duration_value,
                    set_value,
                    note_value,
                )
            try:
                validated = validate_footwork_training_form(form)
                new_record_identity = footwork_record_identity(validated)
                if new_record_identity in existing_record_identities or training_repo.footwork_identity_exists(new_record_identity):
                    error_rows.append(f"第 {idx} 行：重复步法训练记录已跳过")
                    continue
                save_footwork_training_record(form, operator)
                existing_record_identities.add(new_record_identity)
                imported_count += 1
            except ValidationError as exc:
                error_rows.append(f"第 {idx} 行：{exc}")
    except Exception as exc:
        error_rows.append(f"Excel 解析失败：{exc}")
    return imported_count, error_rows


def import_stats_skill_excel(file, operator):
    return import_footwork_training_excel(file, operator)


def build_footwork_import_form(student_no, training_date, footwork_value, duration_value, set_value, note_value):
    athlete = next((player for player in PLAYERS if player["student_no"] == str(student_no).strip()), None)
    if not athlete:
        athlete_id = resolve_athlete_id(student_no, create_if_missing=True)
    else:
        athlete_id = athlete["id"]
    footwork = resolve_dict_by_name(footwork_value, "footwork")
    if not footwork or footwork["dict_code"] == "footwork_root":
        raise ValidationError("步法类型无法识别，请使用系统字典名称。")
    return {
        "athlete_id": athlete_id,
        "training_date": normalize_excel_date(training_date),
        "footwork_dict_id": footwork["id"],
        "duration_minutes": "" if duration_value is None else str(duration_value),
        "set_count": "" if set_value is None else str(set_value),
        "training_note": str(note_value or "").strip(),
    }


def build_member9_footwork_import_form(
    athlete_value,
    training_date,
    footwork_duration,
    hit_score,
    multi_ball_duration,
    intensity_value,
):
    footwork = find_dict_by_code("composite_step")
    note_parts = []
    if hit_score not in (None, ""):
        note_parts.append(f"击球得分：{hit_score}")
    if multi_ball_duration not in (None, ""):
        note_parts.append(f"多球时长：{multi_ball_duration} 分钟")
    if intensity_value not in (None, ""):
        note_parts.append(f"训练强度：{intensity_value}")
    return {
        "athlete_id": resolve_athlete_id(athlete_value, create_if_missing=True),
        "training_date": normalize_excel_date(training_date),
        "footwork_dict_id": footwork["id"],
        "duration_minutes": str(footwork_duration if footwork_duration not in (None, "") else 30),
        "set_count": "4",
        "training_note": "；".join(note_parts),
    }


def build_legacy_footwork_import_form(athlete_value, training_date, footwork_value, duration_value, note_value):
    footwork = resolve_dict_by_name(footwork_value, "footwork")
    if not footwork:
        footwork = find_dict_by_code("composite_step")
    duration = 30 if duration_value is None else duration_value
    note_parts = [str(note_value).strip()] if note_value else []
    if footwork_value and footwork and footwork["dict_name"] != str(footwork_value).strip():
        note_parts.append(f"原始步法字段：{footwork_value}")
    return {
        "athlete_id": resolve_athlete_id(athlete_value, create_if_missing=True),
        "training_date": normalize_excel_date(training_date),
        "footwork_dict_id": footwork["id"],
        "duration_minutes": str(duration),
        "set_count": "4",
        "training_note": "；".join(note_parts),
    }


def parse_optional_float(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def send_workbook(workbook, filename):
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    try:
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except TypeError:
        return send_file(
            buffer,
            as_attachment=True,
            attachment_filename=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def detect_footwork_excel_format(headers):
    header_set = {str(header or "").strip() for header in headers if header}
    if {"运动员学号", "训练日期", "步法类型"} <= header_set:
        return "standard"
    if {"运动员", "训练日期", "步法训练"} <= header_set:
        return "legacy_skill"
    if {"运动员姓名", "训练日期", "步法时长"} <= header_set:
        return "member9"
    if {"运动员名称", "训练日期", "步法时长"} <= header_set:
        return "member9"
    return "standard"


def write_headers(ws, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


def write_training_plan_sheet(ws, plans):
    ws.title = "训练计划"
    headers = ["运动员", "教练", "训练时间", "训练内容", "训练强度", "训练时长(分钟)", "训练地点"]
    write_headers(ws, headers)
    for row_idx, plan in enumerate(plans, 2):
        ws.cell(row=row_idx, column=1, value=plan["athlete_name"])
        ws.cell(row=row_idx, column=2, value=plan["coach_name"])
        ws.cell(row=row_idx, column=3, value=plan["plan_datetime"])
        ws.cell(row=row_idx, column=4, value=plan["content"])
        ws.cell(row=row_idx, column=5, value=plan["intensity"])
        ws.cell(row=row_idx, column=6, value=plan["duration"])
        ws.cell(row=row_idx, column=7, value=plan.get("location", ""))


def write_footwork_sheet(ws, records):
    ws.title = "步法训练记录"
    headers = ["运动员", "训练日期", "步法类型", "训练时长(分钟)", "训练组数", "训练备注"]
    write_headers(ws, headers)
    for row_idx, record in enumerate(records, 2):
        enriched = enrich_footwork_record(record)
        ws.cell(row=row_idx, column=1, value=enriched["athlete_name"])
        ws.cell(row=row_idx, column=2, value=enriched["training_date"])
        ws.cell(row=row_idx, column=3, value=enriched["footwork_type_name"])
        ws.cell(row=row_idx, column=4, value=enriched["duration_minutes"])
        ws.cell(row=row_idx, column=5, value=enriched["set_count"])
        ws.cell(row=row_idx, column=6, value=enriched.get("note", ""))


def write_technique_tactic_sheet(ws, records):
    ws.title = "技战术训练记录"
    headers = [
        "运动员",
        "训练日期",
        "技战术大类",
        "技战术",
        "多球训练球数",
        "发球频率",
        "计划执行率",
        "上台率",
        "落点分布",
        "定性评价",
    ]
    write_headers(ws, headers)
    for row_idx, record in enumerate(records, 2):
        enriched = enrich_technique_tactic_record(record)
        ws.cell(row=row_idx, column=1, value=enriched["athlete_name"])
        ws.cell(row=row_idx, column=2, value=enriched["training_date"])
        ws.cell(row=row_idx, column=3, value=enriched["technique_category_name"])
        ws.cell(row=row_idx, column=4, value=enriched["technique_name"])
        ws.cell(row=row_idx, column=5, value=enriched["multi_ball_count"])
        ws.cell(row=row_idx, column=6, value=enriched["serve_frequency"])
        ws.cell(row=row_idx, column=7, value=enriched["plan_execution_rate"])
        ws.cell(row=row_idx, column=8, value=enriched.get("on_table_rate", ""))
        ws.cell(row=row_idx, column=9, value=enriched.get("landing_distribution", ""))
        ws.cell(row=row_idx, column=10, value=enriched.get("qualitative_comment", ""))


def write_fitness_sheet(ws, tests):
    ws.title = "体能测试记录"
    headers = ["运动员", "测试日期", "测试教练", "上肢力量", "下肢力量", "柔韧性", "耐力", "速度", "综合得分", "备注"]
    write_headers(ws, headers)
    for row_idx, test in enumerate(tests, 2):
        player = find_player(test["athlete_id"])
        coach = next((item for item in COACHES if item["id"] == test["tester_id"]), None)
        ws.cell(row=row_idx, column=1, value=player["name"] if player else "未知运动员")
        ws.cell(row=row_idx, column=2, value=test["test_date"])
        ws.cell(row=row_idx, column=3, value=coach["name"] if coach else "未知教练")
        ws.cell(row=row_idx, column=4, value=test["upper_strength"])
        ws.cell(row=row_idx, column=5, value=test["lower_strength"])
        ws.cell(row=row_idx, column=6, value=test["flexibility"])
        ws.cell(row=row_idx, column=7, value=test["endurance"])
        ws.cell(row=row_idx, column=8, value=test["speed"])
        ws.cell(row=row_idx, column=9, value=test["overall_score"])
        ws.cell(row=row_idx, column=10, value=test.get("notes", ""))


def write_injury_sheet(ws, records):
    ws.title = "伤病记录"
    headers = ["运动员", "伤病日期", "伤病部位", "伤病类型", "严重程度", "诊断说明", "处理方案", "恢复状态", "预计恢复日期", "备注"]
    write_headers(ws, headers)
    row_idx = 2
    for record in records:
        if record.get("is_deleted"):
            continue
        player = find_player(record["athlete_id"])
        ws.cell(row=row_idx, column=1, value=player["name"] if player else "未知运动员")
        ws.cell(row=row_idx, column=2, value=record["injury_date"])
        ws.cell(row=row_idx, column=3, value=record["injury_location"])
        ws.cell(row=row_idx, column=4, value=record["injury_type"])
        ws.cell(row=row_idx, column=5, value=record["severity"])
        ws.cell(row=row_idx, column=6, value=record.get("diagnosis", ""))
        ws.cell(row=row_idx, column=7, value=record.get("treatment", ""))
        ws.cell(row=row_idx, column=8, value=record["recovery_status"])
        ws.cell(row=row_idx, column=9, value=record.get("expected_recovery_date", ""))
        ws.cell(row=row_idx, column=10, value=record.get("notes", ""))
        row_idx += 1


def resolve_athlete_id(value, create_if_missing=False):
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""

    numeric_text = text[:-2] if text.endswith(".0") and text[:-2].isdigit() else text
    if numeric_text.isdigit():
        player = next((item for item in PLAYERS if item["id"] == int(numeric_text)), None)
        if player:
            return str(player["id"])
        player = next(
            (item for item in PLAYERS if item["student_no"] == numeric_text),
            None,
        )
        if player:
            return str(player["id"])

    if text.isdigit():
        player = next((item for item in PLAYERS if item["student_no"] == text), None)
        if player:
            return str(player["id"])

    player = next(
        (item for item in PLAYERS if item["name"] == text or item["student_no"] == text),
        None,
    )
    if player:
        return str(player["id"])
    if create_if_missing and should_create_import_player(text, numeric_text):
        return str(create_import_player_archive(text)["id"])
    raise ValidationError(f"找不到运动员：{text}")


def should_create_import_player(text, numeric_text):
    return bool(text) and not numeric_text.isdigit() and len(text) <= 50


def create_import_player_archive(name):
    player_id = next_id(PLAYERS)
    player = {
        "id": player_id,
        "student_no": next_import_student_no(player_id),
        "name": name,
        "gender": "未录入",
        "age": 18,
        "level": PLAYER_LEVEL_LABELS["youth"],
        "skill_level": PLAYER_LEVEL_LABELS["youth"],
        "level_code": "youth",
        "play_style": "Excel导入待补充",
        "grip": "",
        "contact_phone": "",
        "injury_status": PLAYER_INJURY_STATUS_LABELS["healthy"],
        "injury_status_code": "healthy",
        "source": "excel_import",
    }
    PLAYERS.append(player)
    return player


def next_import_student_no(player_id):
    candidate_number = player_id
    while True:
        student_no = f"IMP{candidate_number:04d}"
        if not any(item["student_no"] == student_no for item in PLAYERS):
            return student_no
        candidate_number += 1


def resolve_option_code(value, options):
    if value is None:
        return ""
    text = str(value).strip()
    if text in options:
        return text
    for code, label in options.items():
        if text == label:
            return code
    return text


def normalize_excel_date(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=int(value))).strftime("%Y-%m-%d")
    return str(value).strip()[:10]


def validate_training_plan_data(athlete_id, coach_id, plan_datetime, content, intensity, duration):
    """校验训练计划数据，返回错误列表"""
    errors = []
    
    if not all([athlete_id, coach_id, plan_datetime, content]):
        errors.append("请填写完整信息（运动员、教练、日期、内容为必填）")
        return errors


    athlete = next((p for p in PLAYERS if str(p["id"]) == athlete_id), None)
    coach = next((c for c in COACHES if str(c["id"]) == coach_id), None)
    if not athlete:
        errors.append("所选运动员不存在，请重新选择")
    if not coach:
        errors.append("所选教练不存在，请重新选择")
    if not athlete or not coach:
        return errors

    from datetime import datetime
    try:
        # 提取日期部分（前10位）
        date_part = plan_datetime[:10] if plan_datetime else ""
        if not date_part:
            errors.append("日期格式错误")
        else:
            selected_date = datetime.strptime(date_part, "%Y-%m-%d")
            if selected_date < datetime.now():
                errors.append("训练日期不能早于今天")
    except ValueError:
        errors.append("日期格式错误，请使用 YYYY-MM-DD 格式")

    valid_intensities = ["低", "中", "高", "极高"]
    if intensity not in valid_intensities:
        errors.append("训练强度必须为：低、中、高、极高")

    try:
        duration_int = int(duration) if duration else 60
        if duration_int <= 0:
            errors.append("训练时长必须大于 0 分钟")
        elif duration_int > 600:
            errors.append("训练时长不能超过 600 分钟（10小时）")
    except (ValueError, TypeError):
        errors.append("训练时长必须为有效的数字")

    if len(content) > 500:
        errors.append("训练内容不能超过 500 个字符")

    return errors

def filter_injury_records(args):
    predicates = []
    query_errors = []
    player_keyword = args.get("player_keyword", "").strip().lower()
    location_keyword = args.get("location_keyword", "").strip().lower()
    severity = args.get("severity", "").strip()
    recovery_status = args.get("recovery_status", "").strip()
    date_from = parse_optional_query_date(args.get("date_from", "").strip(), "开始日期", query_errors)
    date_to = parse_optional_query_date(args.get("date_to", "").strip(), "结束日期", query_errors)
    active_only = args.get("active_only", "").strip()
    show_archived = args.get("show_archived", "").strip()

    if player_keyword:
        predicates.append(
            lambda record, value=player_keyword: value in record["player_name"].lower()
            or value in record["student_no"].lower()
        )
    if location_keyword:
        predicates.append(
            lambda record, value=location_keyword: value in record["injury_location"].lower()
            or value in record["injury_type"].lower()
        )
    if severity:
        if severity in INJURY_SEVERITY_OPTIONS:
            predicates.append(lambda record, value=severity: record["severity"] == value)
        else:
            query_errors.append("伤病程度筛选条件非法，已忽略该条件。")
    if recovery_status:
        if recovery_status in INJURY_RECOVERY_STATUS_OPTIONS:
            predicates.append(lambda record, value=recovery_status: record["recovery_status"] == value)
        else:
            query_errors.append("恢复状态筛选条件非法，已忽略该条件。")
    if date_from and date_to and date_from > date_to:
        query_errors.append("开始日期不能晚于结束日期，已忽略日期范围条件。")
        date_from = ""
        date_to = ""
    if date_from:
        predicates.append(lambda record, value=date_from: record["injury_date"] >= value)
    if date_to:
        predicates.append(lambda record, value=date_to: record["injury_date"] <= value)
    if active_only:
        predicates.append(lambda record: record["recovery_status"] in {"治疗中", "康复中"})

    source_records = INJURY_RECORDS if show_archived else [
        item for item in INJURY_RECORDS if not item.get("is_deleted")
    ]
    records = [enrich_injury_record(item) for item in source_records]
    records.sort(key=lambda item: (item["injury_date"], item["id"]), reverse=True)
    if not predicates:
        return records, 0, query_errors
    return [record for record in records if all(check(record) for check in predicates)], len(predicates), query_errors


def enrich_injury_record(record):
    player = next((item for item in PLAYERS if item["id"] == record["athlete_id"]), None)
    expected_date = record.get("expected_recovery_date") or ""
    overdue = False
    if expected_date and record["recovery_status"] != "已恢复":
        overdue = expected_date < datetime.now().strftime("%Y-%m-%d")
    base = dict(record)
    base.update(
        {
            "player_name": player["name"] if player else "未知运动员",
            "student_no": player["student_no"] if player else "-",
            "level": (player.get("skill_level") or player.get("level")) if player else "-",
            "player_injury_status": player["injury_status"] if player else "-",
            "player_injury_status_code": player["injury_status_code"] if player else "healthy",
            "severity_class": injury_severity_class(record["severity"]),
            "recovery_class": injury_recovery_class(record["recovery_status"]),
            "is_active": record["recovery_status"] in {"治疗中", "康复中"},
            "is_overdue": overdue,
            "followup_count": count_followups(record["id"]),
            "latest_followup": latest_followup(record["id"]),
            "training_alert": build_training_alert_for_record(record),
        }
    )
    return base


def injury_severity_class(severity):
    return {
        "轻微": "success",
        "中度": "warning",
        "严重": "danger",
    }.get(severity, "secondary")


def injury_recovery_class(recovery_status):
    return {
        "治疗中": "danger",
        "康复中": "info",
        "已恢复": "success",
    }.get(recovery_status, "secondary")


def build_injury_summary(records):
    status_counts = {status: 0 for status in INJURY_RECOVERY_STATUS_OPTIONS}
    location_counts = {}
    active_count = 0
    overdue_count = 0
    serious_count = 0
    for record in records:
        status_counts[record["recovery_status"]] += 1
        location_counts[record["injury_location"]] = location_counts.get(record["injury_location"], 0) + 1
        if record["recovery_status"] in {"治疗中", "康复中"}:
            active_count += 1
        if record["is_overdue"]:
            overdue_count += 1
        if record["severity"] == "严重":
            serious_count += 1

    top_locations = sorted(location_counts.items(), key=lambda item: item[1], reverse=True)[:5]
    return {
        "record_count": len(records),
        "active_count": active_count,
        "treating_count": status_counts["治疗中"],
        "rehab_count": status_counts["康复中"],
        "recovered_count": status_counts["已恢复"],
        "overdue_count": overdue_count,
        "serious_count": serious_count,
        "status_pie": [{"name": key, "value": value} for key, value in status_counts.items()],
        "location_names": [item[0] for item in top_locations],
        "location_counts": [item[1] for item in top_locations],
    }


def get_editing_injury_record(edit_id):
    if not edit_id.isdigit():
        return None
    record = next(
        (item for item in INJURY_RECORDS if item["id"] == int(edit_id) and not item.get("is_deleted")),
        None,
    )
    return enrich_injury_record(record) if record else None


def save_injury_record(form, user):
    validated = validate_injury_form(form)
    enforce_injury_write_permission(validated, user)
    original_records = deepcopy(INJURY_RECORDS)
    original_players = deepcopy(PLAYERS)
    affected_athlete_ids = {validated["athlete_id"]}
    try:
        record_id = validated.pop("record_id")
        if record_id:
            target = next(
                (item for item in INJURY_RECORDS if item["id"] == record_id and not item.get("is_deleted")),
                None,
            )
            if not target:
                raise ValidationError("要修改的伤病记录不存在。")
            affected_athlete_ids.add(target["athlete_id"])
            target.update(validated)
            target["created_by"] = user["username"]
        else:
            INJURY_RECORDS.append(
                {
                    "id": next_id(INJURY_RECORDS),
                    **validated,
                    "created_by": user["username"],
                    "is_deleted": False,
                    "deleted_by": "",
                    "deleted_at": "",
                    "delete_reason": "",
                }
            )

        for athlete_id in affected_athlete_ids:
            refresh_athlete_injury_status(athlete_id)
    except Exception:
        INJURY_RECORDS[:] = original_records
        PLAYERS[:] = original_players
        raise


def archive_injury_record(form, user):
    if user["role"] != "admin":
        raise ValidationError("只有管理员可以作废归档伤病记录。")
    record_id = parse_int_field(form.get("archive_record_id", "").strip(), "伤病记录")
    reason = form.get("delete_reason", "").strip()
    if not reason:
        raise ValidationError("作废原因不能为空。")
    if len(reason) > 120:
        raise ValidationError("作废原因不能超过 120 个字符。")

    original_records = deepcopy(INJURY_RECORDS)
    original_players = deepcopy(PLAYERS)
    try:
        target = next(
            (item for item in INJURY_RECORDS if item["id"] == record_id and not item.get("is_deleted")),
            None,
        )
        if not target:
            raise ValidationError("要作废的伤病记录不存在或已归档。")
        target["is_deleted"] = True
        target["deleted_by"] = user["username"]
        target["deleted_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        target["delete_reason"] = reason
        refresh_athlete_injury_status(target["athlete_id"])
    except Exception:
        INJURY_RECORDS[:] = original_records
        PLAYERS[:] = original_players
        raise


def save_injury_followup(form, user):
    record_id = parse_int_field(form.get("followup_record_id", "").strip(), "伤病记录")
    target = next(
        (item for item in INJURY_RECORDS if item["id"] == record_id and not item.get("is_deleted")),
        None,
    )
    if not target:
        raise ValidationError("要跟踪的伤病记录不存在或已归档。")

    followup_date = parse_date_field(form.get("followup_date", "").strip(), "复诊日期")
    if followup_date < target["injury_date"]:
        raise ValidationError("复诊日期不能早于伤病日期。")
    pain_score = parse_int_range(form.get("pain_score", "").strip(), "疼痛评分", 0, 10)
    training_limit = form.get("training_limit", "").strip()
    advice = form.get("advice", "").strip()
    reviewer = form.get("reviewer", "").strip()

    for field_label, value, max_length in (
        ("训练限制", training_limit, 160),
        ("复诊建议", advice, 160),
        ("复诊人", reviewer, 30),
    ):
        if not value:
            raise ValidationError(f"{field_label}不能为空。")
        if len(value) > max_length:
            raise ValidationError(f"{field_label}不能超过 {max_length} 个字符。")

    INJURY_FOLLOWUPS.append(
        {
            "id": next_id(INJURY_FOLLOWUPS),
            "injury_record_id": record_id,
            "followup_date": followup_date,
            "pain_score": pain_score,
            "training_limit": training_limit,
            "advice": advice,
            "reviewer": reviewer,
            "created_by": user["username"],
        }
    )


def validate_injury_form(form):
    record_id = form.get("record_id", "").strip()
    if record_id and not record_id.isdigit():
        raise ValidationError("伤病记录编号非法，请从列表中选择要修改的记录。")
    athlete_id = parse_int_field(form.get("athlete_id", "").strip(), "运动员")
    if not any(player["id"] == athlete_id for player in PLAYERS):
        raise ValidationError("所选运动员不存在，请重新选择。")

    injury_date = parse_date_field(form.get("injury_date", "").strip(), "伤病日期")
    expected_recovery_date = form.get("expected_recovery_date", "").strip()
    if expected_recovery_date:
        expected_recovery_date = parse_date_field(expected_recovery_date, "预计恢复日期")
        if expected_recovery_date < injury_date:
            raise ValidationError("预计恢复日期不能早于伤病日期。")

    injury_location = form.get("injury_location", "").strip()
    if not injury_location:
        raise ValidationError("伤病部位不能为空。")
    if len(injury_location) > 50:
        raise ValidationError("伤病部位不能超过 50 个字符。")

    injury_type = form.get("injury_type", "").strip()
    if not injury_type:
        raise ValidationError("伤病类型不能为空。")
    if len(injury_type) > 50:
        raise ValidationError("伤病类型不能超过 50 个字符。")

    severity = form.get("severity", "").strip()
    if severity not in INJURY_SEVERITY_OPTIONS:
        raise ValidationError("伤病程度非法，请从轻微、中度、严重中选择。")

    recovery_status = form.get("recovery_status", "").strip()
    if recovery_status not in INJURY_RECOVERY_STATUS_OPTIONS:
        raise ValidationError("恢复状态非法，请从治疗中、康复中、已恢复中选择。")

    diagnosis = form.get("diagnosis", "").strip()
    treatment = form.get("treatment", "").strip()
    notes = form.get("notes", "").strip()
    for field_label, value, max_length in (
        ("诊断说明", diagnosis, 180),
        ("处理方案", treatment, 180),
        ("备注", notes, 120),
    ):
        if len(value) > max_length:
            raise ValidationError(f"{field_label}不能超过 {max_length} 个字符。")

    if recovery_status == "已恢复" and not expected_recovery_date:
        expected_recovery_date = injury_date

    return {
        "record_id": int(record_id) if record_id.isdigit() else None,
        "athlete_id": athlete_id,
        "injury_date": injury_date,
        "injury_location": injury_location,
        "injury_type": injury_type,
        "severity": severity,
        "diagnosis": diagnosis,
        "treatment": treatment,
        "recovery_status": recovery_status,
        "expected_recovery_date": expected_recovery_date,
        "notes": notes,
    }


def enforce_injury_write_permission(validated, user):
    if user["role"] == "admin":
        return
    if validated["record_id"]:
        raise ValidationError("普通教练只能新增伤病记录，修改记录需管理员处理。")
    if validated["severity"] == "严重":
        raise ValidationError("严重伤病需管理员确认后登记。")
    if validated["recovery_status"] == "已恢复":
        raise ValidationError("恢复完成状态需管理员确认。")


def refresh_athlete_injury_status(athlete_id):
    player = next((item for item in PLAYERS if item["id"] == athlete_id), None)
    if not player:
        raise RuntimeError("无法刷新健康状态：运动员不存在。")

    active_records = [
        record for record in INJURY_RECORDS
        if record["athlete_id"] == athlete_id
        and not record.get("is_deleted")
        and record["recovery_status"] in {"治疗中", "康复中"}
    ]
    if any(record["recovery_status"] == "治疗中" and record["severity"] == "严重" for record in active_records):
        status = "伤病中"
    elif any(record["recovery_status"] == "治疗中" for record in active_records):
        status = "观察中"
    elif any(record["recovery_status"] == "康复中" for record in active_records):
        status = "康复中"
    else:
        status = "健康"

    status_meta = ATHLETE_INJURY_STATUS_META[status]
    player["injury_status"] = status
    player["injury_status_code"] = status_meta["code"]


def build_injury_redirect_query(form):
    edit_id = form.get("record_id", "").strip()
    if edit_id:
        return {"edit_id": edit_id}
    return {}


def group_followups_by_record():
    grouped = {}
    for followup in sorted(
        INJURY_FOLLOWUPS,
        key=lambda item: (item["followup_date"], item["id"]),
        reverse=True,
    ):
        grouped.setdefault(followup["injury_record_id"], []).append(followup)
    return grouped


def count_followups(record_id):
    return sum(1 for item in INJURY_FOLLOWUPS if item["injury_record_id"] == record_id)


def latest_followup(record_id):
    items = [
        item for item in INJURY_FOLLOWUPS
        if item["injury_record_id"] == record_id
    ]
    if not items:
        return None
    return sorted(items, key=lambda item: (item["followup_date"], item["id"]), reverse=True)[0]


def build_training_alert_for_record(record):
    if record.get("is_deleted") or record["recovery_status"] == "已恢复":
        return ""
    if record["severity"] == "严重" and record["recovery_status"] == "治疗中":
        return "禁止高强度和对抗训练，需调整为康复或休训方案。"
    if record["recovery_status"] == "治疗中":
        return "避免高负荷专项训练，训练计划需降低强度。"
    if record["recovery_status"] == "康复中":
        return "仅安排低到中等强度过渡训练，复训前保留评估。"
    return ""


def build_training_alerts(records):
    alerts = []
    for record in records:
        alert = record.get("training_alert") or build_training_alert_for_record(record)
        if alert:
            alerts.append(
                {
                    "record_id": record["id"],
                    "player_name": record["player_name"],
                    "injury_location": record["injury_location"],
                    "severity": record["severity"],
                    "recovery_status": record["recovery_status"],
                    "alert": alert,
                }
            )
    return alerts[:6]



def filter_fitness_tests(args):
    predicates = []
    player_keyword = args.get("player_keyword", "").strip().lower()
    date_from = args.get("date_from", "").strip()
    date_to = args.get("date_to", "").strip()
    risk_level = args.get("risk_level", "").strip()
    intensity = args.get("intensity", "").strip()
    score_min = args.get("score_min", "").strip()
    lower_strength_min = args.get("lower_strength_min", "").strip()
    speed_min = args.get("speed_min", "").strip()

    if player_keyword:
        predicates.append(
            lambda record, value=player_keyword: value in record["player_name"].lower()
            or value in record["student_no"].lower()
        )
    if date_from:
        predicates.append(lambda record, value=date_from: record["test_date"] >= value)
    if date_to:
        predicates.append(lambda record, value=date_to: record["test_date"] <= value)
    if risk_level:
        predicates.append(lambda record, value=risk_level: record["risk_code"] == value)
    if intensity:
        predicates.append(lambda record, value=intensity: record["plan_intensity"] == value)
    if is_float_value(score_min):
        predicates.append(lambda record, value=float(score_min): record["overall_score"] >= value)
    if is_float_value(lower_strength_min):
        predicates.append(lambda record, value=float(lower_strength_min): record["lower_strength"] >= value)
    if is_float_value(speed_min):
        predicates.append(lambda record, value=float(speed_min): record["speed"] >= value)

    records = [enrich_fitness_record(item) for item in FITNESS_TESTS]
    records.sort(key=lambda item: (item["test_date"], item["id"]), reverse=True)
    if not predicates:
        return records, 0
    filtered = [record for record in records if all(check(record) for check in predicates)]
    return filtered, len(predicates)


def enrich_fitness_record(record):
    player = next((item for item in PLAYERS if item["id"] == record["athlete_id"]), None)
    coach = next((item for item in COACHES if item["id"] == record["tester_id"]), None)
    sync_plan = next((item for item in TRAINING_SYNC_LOGS if item["fitness_test_id"] == record["id"]), None)
    risk = evaluate_fitness_risk(record)
    score = calculate_fitness_score(record)
    upper_strength_status = classify_metric_status(record["upper_strength"], 70, 80, lower_is_worse=True)
    lower_strength_status = classify_metric_status(record["lower_strength"], 70, 80, lower_is_worse=True)
    flexibility_status = classify_metric_status(record["flexibility"], 70, 80, lower_is_worse=True)
    endurance_status = classify_metric_status(record["endurance"], 75, 85, lower_is_worse=True)
    speed_status = classify_metric_status(record["speed"], 75, 85, lower_is_worse=True)
    base = dict(record)
    base.update(
        {
            "player_name": player["name"] if player else "未知运动员",
            "student_no": player["student_no"] if player else "-",
            "level": (player.get("skill_level") or player.get("level")) if player else "-",
            "tester_name": coach["name"] if coach else "未指定",
            "risk_code": risk["code"],
            "risk_label": risk["label"],
            "risk_class": risk["class"],
            "fitness_score": score,
            "upper_strength_status": upper_strength_status,
            "lower_strength_status": lower_strength_status,
            "flexibility_status": flexibility_status,
            "endurance_status": endurance_status,
            "speed_status": speed_status,
            "plan_name": sync_plan["plan_name"] if sync_plan else "-",
            "plan_hours": sync_plan["hours"] if sync_plan else 0,
            "plan_intensity": sync_plan["intensity"] if sync_plan else "",
            "plan_status": sync_plan["status"] if sync_plan else "-",
        }
    )
    return base


def evaluate_fitness_risk(record):
    alerts = 0
    observes = 0
    if record["upper_strength"] < 70:
        alerts += 1
    elif record["upper_strength"] < 80:
        observes += 1
    if record["lower_strength"] < 70:
        alerts += 1
    elif record["lower_strength"] < 80:
        observes += 1
    if record["flexibility"] < 70:
        alerts += 1
    elif record["flexibility"] < 80:
        observes += 1
    if record["endurance"] < 75:
        alerts += 1
    elif record["endurance"] < 85:
        observes += 1
    if record["speed"] < 75:
        alerts += 1
    elif record["speed"] < 85:
        observes += 1
    if alerts >= 2:
        return {"code": "alert", "label": "预警", "class": "danger"}
    if alerts == 1 or observes >= 2:
        return {"code": "observe", "label": "观察", "class": "warning"}
    return {"code": "stable", "label": "稳定", "class": "success"}


def calculate_fitness_score(record):
    score = (
        record["upper_strength"]
        + record["lower_strength"]
        + record["flexibility"]
        + record["endurance"]
        + record["speed"]
    ) / 5
    return round(score, 1)


def classify_metric_status(value, alert_threshold, observe_threshold, *, lower_is_worse):
    if lower_is_worse:
        if value < alert_threshold:
            return "alert"
        if value < observe_threshold:
            return "observe"
        return "normal"
    if value > alert_threshold:
        return "alert"
    if value > observe_threshold:
        return "observe"
    return "normal"


def build_fitness_summary(records):
    risk_counts = {"稳定": 0, "观察": 0, "预警": 0}
    monthly_map = {}
    player_map = {}
    for record in records:
        risk_counts[record["risk_label"]] += 1
        month_key = record["test_date"][:7]
        monthly_stats = monthly_map.setdefault(
            month_key,
            {"score_total": 0.0, "speed_total": 0.0, "hours_total": 0.0, "count": 0},
        )
        monthly_stats["score_total"] += record["fitness_score"]
        monthly_stats["speed_total"] += record["speed"]
        monthly_stats["hours_total"] += record["plan_hours"]
        monthly_stats["count"] += 1

        player_stats = player_map.setdefault(record["player_name"], {"score_total": 0.0, "count": 0})
        player_stats["score_total"] += record["fitness_score"]
        player_stats["count"] += 1

    month_labels = sorted(monthly_map.keys())
    monthly_scores = [round(monthly_map[key]["score_total"] / monthly_map[key]["count"], 1) for key in month_labels]
    monthly_speed = [round(monthly_map[key]["speed_total"] / monthly_map[key]["count"], 1) for key in month_labels]
    monthly_hours = [round(monthly_map[key]["hours_total"], 1) for key in month_labels]

    player_scores = sorted(
        (
            {
                "name": name,
                "score": round(stats["score_total"] / stats["count"], 1),
            }
            for name, stats in player_map.items()
        ),
        key=lambda item: item["score"],
        reverse=True,
    )
    top_player_scores = player_scores[:5]

    return {
        "record_count": len(records),
        "warning_count": risk_counts["预警"],
        "average_score": round(sum(record["fitness_score"] for record in records) / len(records), 1) if records else 0,
        "avg_speed": round(sum(record["speed"] for record in records) / len(records), 1) if records else 0,
        "risk_pie": [{"name": key, "value": value} for key, value in risk_counts.items()],
        "month_labels": month_labels,
        "monthly_scores": monthly_scores,
        "monthly_speed": monthly_speed,
        "monthly_hours": monthly_hours,
        "player_names": [item["name"] for item in top_player_scores],
        "player_scores": [item["score"] for item in top_player_scores],
    }


def get_editing_fitness_record(edit_id):
    if not edit_id.isdigit():
        return None
    record = next((item for item in FITNESS_TESTS if item["id"] == int(edit_id)), None)
    return enrich_fitness_record(record) if record else None


def save_fitness_test(form, operator):
    validated = validate_fitness_form(form)
    original_tests = deepcopy(FITNESS_TESTS)
    original_logs = deepcopy(TRAINING_SYNC_LOGS)
    try:
        record_id = validated.pop("record_id")
        if record_id:
            target = next((item for item in FITNESS_TESTS if item["id"] == record_id), None)
            if not target:
                raise ValidationError("要修改的体能测试记录不存在。")
            target.update(validated)
            target["created_by"] = operator
            sync_log = next((item for item in TRAINING_SYNC_LOGS if item["fitness_test_id"] == record_id), None)
            if sync_log:
                sync_log.update(
                    {
                        "athlete_id": validated["athlete_id"],
                        "coach_id": validated["tester_id"],
                        "sync_date": validated["test_date"],
                        "plan_name": validated["plan_name"],
                        "hours": validated["hours"],
                        "intensity": validated["intensity"],
                        "status": validated["plan_status"],
                    }
                )
        else:
            new_id = next_id(FITNESS_TESTS)
            FITNESS_TESTS.append(
                {
                    "id": new_id,
                    **validated,
                    "created_by": operator,
                }
            )
            TRAINING_SYNC_LOGS.append(
                {
                    "id": next_id(TRAINING_SYNC_LOGS),
                    "fitness_test_id": new_id,
                    "athlete_id": validated["athlete_id"],
                    "coach_id": validated["tester_id"],
                    "sync_date": validated["test_date"],
                    "plan_name": validated["plan_name"],
                    "hours": validated["hours"],
                    "intensity": validated["intensity"],
                    "status": validated["plan_status"],
                }
            )
    except Exception:
        FITNESS_TESTS[:] = original_tests
        TRAINING_SYNC_LOGS[:] = original_logs
        raise


def validate_fitness_form(form):
    record_id = form.get("record_id", "").strip()
    athlete_id = parse_int_field(form.get("athlete_id", "").strip(), "运动员")
    if not any(player["id"] == athlete_id for player in PLAYERS):
        raise ValidationError("所选运动员不存在，请重新选择。")
    tester_id = parse_int_field(form.get("tester_id", "").strip(), "测试教练")
    if not any(coach["id"] == tester_id for coach in COACHES):
        raise ValidationError("所选测试教练不存在，请重新选择。")

    test_date = form.get("test_date", "").strip()
    try:
        datetime.strptime(test_date, "%Y-%m-%d")
    except ValueError as exc:
        raise ValidationError("测试日期格式错误，请使用 YYYY-MM-DD。") from exc

    upper_strength = parse_float_range(form.get("upper_strength", "").strip(), "上肢力量", 0, 100)
    lower_strength = parse_float_range(form.get("lower_strength", "").strip(), "下肢力量", 0, 100)
    flexibility = parse_float_range(form.get("flexibility", "").strip(), "柔韧性", 0, 100)
    endurance = parse_float_range(form.get("endurance", "").strip(), "耐力", 0, 100)
    speed = parse_float_range(form.get("speed", "").strip(), "速度", 0, 100)
    hours = parse_float_range(form.get("hours", "").strip(), "训练时长", 0, 999.9)

    intensity = form.get("intensity", "").strip()
    if intensity not in INTENSITY_LABELS:
        raise ValidationError("训练强度非法，请从低、中、高、极高中选择。")
    plan_status = form.get("plan_status", "").strip()
    if plan_status not in {"进行中", "已完成", "已取消"}:
        raise ValidationError("训练计划状态非法。")
    plan_name = form.get("plan_name", "").strip()
    if not plan_name:
        raise ValidationError("训练计划名称不能为空。")

    notes = form.get("notes", "").strip()
    if len(notes) > 120:
        raise ValidationError("备注不能超过 120 个字符。")

    overall_score = round((upper_strength + lower_strength + flexibility + endurance + speed) / 5, 2)

    return {
        "record_id": int(record_id) if record_id.isdigit() else None,
        "athlete_id": athlete_id,
        "test_date": test_date,
        "tester_id": tester_id,
        "upper_strength": upper_strength,
        "lower_strength": lower_strength,
        "flexibility": flexibility,
        "endurance": endurance,
        "speed": speed,
        "overall_score": overall_score,
        "plan_name": plan_name,
        "hours": hours,
        "intensity": intensity,
        "plan_status": plan_status,
        "notes": notes,
    }


def build_redirect_query(form):
    edit_id = form.get("record_id", "").strip()
    if edit_id:
        return {"edit_id": edit_id}
    return {}


def next_id(rows):
    return max((item["id"] for item in rows), default=0) + 1


def parse_int_field(value, field_name):
    if not value:
        raise ValidationError(f"{field_name}不能为空。")
    try:
        return int(value)
    except ValueError as exc:
        raise ValidationError(f"{field_name}必须为整数。") from exc


def parse_int_range(value, field_name, min_value, max_value):
    parsed = parse_int_field(value, field_name)
    if parsed < min_value or parsed > max_value:
        raise ValidationError(f"{field_name}必须介于 {min_value} 和 {max_value} 之间。")
    return parsed


def parse_float_range(value, field_name, min_value, max_value):
    if not value:
        raise ValidationError(f"{field_name}不能为空。")
    try:
        parsed = float(value)
    except ValueError as exc:
        raise ValidationError(f"{field_name}必须为数字类型。") from exc
    if parsed < min_value or parsed > max_value:
        raise ValidationError(f"{field_name}必须介于 {min_value} 和 {max_value} 之间。")
    return round(parsed, 3)


def parse_date_field(value, field_name):
    if not value:
        raise ValidationError(f"{field_name}不能为空。")
    try:
        datetime.strptime(value, "%Y-%m-%d")
    except ValueError as exc:
        raise ValidationError(f"{field_name}格式错误，请使用 YYYY-MM-DD。") from exc
    return value


def parse_optional_query_date(value, field_name, errors):
    if not value:
        return ""
    try:
        datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        errors.append(f"{field_name}格式错误，已忽略该条件。")
        return ""
    return value


def is_float_value(value):
    if not value:
        return False
    try:
        float(value)
        return True
    except ValueError:
        return False


def module_page(module_name, module_desc):
    return render_template(
        "module_overview.html",
        module_name=module_name,
        module_desc=module_desc,
        features=MODULE_FEATURES.get(module_name, []),
    )


app = create_app()


def _should_use_reloader():
    override = os.getenv("FLASK_USE_RELOADER", "").strip().lower()
    if override in {"0", "false", "no"}:
        return False
    if override in {"1", "true", "yes"}:
        return True
    # VS Code debugpy + Flask reloader triggers WinError 10038 on Windows.
    if any("debugpy" in str(arg) for arg in sys.argv):
        return False
    return True


if __name__ == "__main__":
    app.run(debug=True, use_reloader=_should_use_reloader())
